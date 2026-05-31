from dataclasses import dataclass
from datetime import datetime, time, timedelta
from decimal import Decimal
from io import BytesIO
import unicodedata

from django.db import transaction
from django.db.models import Max
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import ActivityLog, Position, TaskTemplate


ACTIVITY_IMPORT_HEADERS = [
    'tipo_usuario_codigo',
    'tipo_usuario_nome',
    'dia_semana',
    'inicio',
    'fim',
    'descricao',
    'resultado_esperado',
    'recorrencia',
    'evidencia_exigida',
    'onde_comprovar',
    'observacoes',
    'ativo',
]

REQUIRED_VALUE_HEADERS = {
    'dia_semana',
    'inicio',
    'fim',
    'descricao',
    'resultado_esperado',
    'recorrencia',
}

WEEKDAY_MAP = {
    'segunda': 0,
    'segunda feira': 0,
    '2': 0,
    '2 feira': 0,
    'terca': 1,
    'terca feira': 1,
    'terça': 1,
    'terça feira': 1,
    '3': 1,
    '3 feira': 1,
    'quarta': 2,
    'quarta feira': 2,
    '4': 2,
    '4 feira': 2,
    'quinta': 3,
    'quinta feira': 3,
    '5': 3,
    '5 feira': 3,
    'sexta': 4,
    'sexta feira': 4,
    '6': 4,
    '6 feira': 4,
    'sabado': 5,
    'sábado': 5,
}

FREQUENCY_MAP = {
    'diaria': TaskTemplate.FREQ_DAILY,
    'diario': TaskTemplate.FREQ_DAILY,
    'daily': TaskTemplate.FREQ_DAILY,
    'semanal': TaskTemplate.FREQ_WEEKLY,
    'weekly': TaskTemplate.FREQ_WEEKLY,
    'mensal': TaskTemplate.FREQ_MONTHLY,
    'monthly': TaskTemplate.FREQ_MONTHLY,
    'anual': TaskTemplate.FREQ_ANNUAL,
    'annual': TaskTemplate.FREQ_ANNUAL,
}

TRUE_VALUES = {'1', 's', 'sim', 'true', 'verdadeiro', 'ativo', 'ativa', 'yes'}
FALSE_VALUES = {'0', 'n', 'nao', 'não', 'false', 'falso', 'inativo', 'inativa', 'no'}


@dataclass
class ActivityImportRow:
    excel_row: int
    position: Position
    day_of_week: int
    start_time: time
    end_time: time
    title: str
    expected_result: str
    frequency: str
    evidence_required: str
    proof_location: str
    notes: str
    active: bool


@dataclass
class ActivityImportResult:
    created: int
    updated: int
    processed: int


def _normalize(value):
    text = str(value or '').strip().lower()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(char for char in text if not unicodedata.combining(char))
    return text.replace('_', ' ').replace('-', ' ')


def _normalize_header(value):
    text = str(value or '').strip().lower()
    return text.replace(' ', '_').replace('-', '_')


def _clean_text(value):
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'Sim' if value else 'Não'
    text = str(value).strip()
    if text.endswith('.0') and isinstance(value, (int, float, Decimal)):
        text = text[:-2]
    return text


def _row_value(row, header_map, header):
    index = header_map.get(header)
    if not index:
        return None
    return row[index - 1]


def _parse_time(value):
    if value is None or value == '':
        return None
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, timedelta):
        total_seconds = int(value.total_seconds())
        if 0 <= total_seconds < 86400:
            return (datetime.min + value).time().replace(second=0, microsecond=0)
        return None
    if isinstance(value, (int, float, Decimal)) and 0 <= float(value) < 1:
        total_seconds = round(float(value) * 86400)
        return time(total_seconds // 3600, (total_seconds % 3600) // 60)

    text = _clean_text(value)
    for fmt in ('%H:%M', '%H:%M:%S'):
        try:
            return datetime.strptime(text, fmt).time().replace(second=0, microsecond=0)
        except ValueError:
            pass
    return None


def _parse_weekday(value):
    key = _normalize(value)
    return WEEKDAY_MAP.get(key)


def _parse_frequency(value):
    key = _normalize(value)
    return FREQUENCY_MAP.get(key)


def _parse_active(value):
    if value is None or value == '':
        return True
    key = _normalize(value)
    if key in TRUE_VALUES:
        return True
    if key in FALSE_VALUES:
        return False
    return None


def _validate_text(label, value, max_length, errors, required=False):
    text = _clean_text(value)
    if required and not text:
        errors.append(f'{label} é obrigatório.')
    if text and len(text) > max_length:
        errors.append(f'{label} deve ter no máximo {max_length} caracteres.')
    return text


def _resolve_position(code_value, name_value, errors):
    code = _clean_text(code_value)
    name = _clean_text(name_value)
    if not code and not name:
        errors.append('Informe tipo_usuario_codigo ou tipo_usuario_nome.')
        return None

    by_code = Position.objects.filter(code__iexact=code).first() if code else None
    by_name = Position.objects.filter(name__iexact=name).first() if name else None
    if code and not by_code:
        errors.append(f'Tipo/cargo com código "{code}" não encontrado.')
    if name and not by_name:
        errors.append(f'Tipo/cargo com nome "{name}" não encontrado.')
    if by_code and by_name and by_code.pk != by_name.pk:
        errors.append('tipo_usuario_codigo e tipo_usuario_nome apontam para tipos/cargos diferentes.')
        return None
    return by_code or by_name


def _validate_headers(header_map):
    errors = []
    if not {'tipo_usuario_codigo', 'tipo_usuario_nome'} & set(header_map):
        errors.append('Inclua a coluna tipo_usuario_codigo ou tipo_usuario_nome.')
    for header in sorted(REQUIRED_VALUE_HEADERS):
        if header not in header_map:
            errors.append(f'Cabeçalho obrigatório ausente: {header}.')
    return errors


def parse_activity_import(uploaded_file):
    try:
        workbook = load_workbook(uploaded_file, data_only=True, read_only=True)
    except Exception:
        return [], [{'row': 'Arquivo', 'errors': ['Não foi possível abrir o arquivo. Envie um XLSX válido.']}]

    worksheet = workbook['Atividades'] if 'Atividades' in workbook.sheetnames else workbook.active
    raw_headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
    header_map = {}
    for index, value in enumerate(raw_headers, start=1):
        header = _normalize_header(value)
        if header:
            header_map[header] = index

    header_errors = _validate_headers(header_map)
    if header_errors:
        return [], [{'row': 'Cabeçalho', 'errors': header_errors}]

    parsed_rows = []
    errors_by_row = []
    seen_keys = {}

    for excel_row, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(_clean_text(value) for value in row):
            continue

        errors = []
        position = _resolve_position(
            _row_value(row, header_map, 'tipo_usuario_codigo'),
            _row_value(row, header_map, 'tipo_usuario_nome'),
            errors,
        )

        day_of_week = _parse_weekday(_row_value(row, header_map, 'dia_semana'))
        if day_of_week is None:
            errors.append('dia_semana inválido. Use Segunda-feira, Terça-feira, Quarta-feira, Quinta-feira, Sexta-feira ou Sábado.')

        start_time = _parse_time(_row_value(row, header_map, 'inicio'))
        if not start_time:
            errors.append('inicio inválido. Use horário no formato HH:MM.')

        end_time = _parse_time(_row_value(row, header_map, 'fim'))
        if not end_time:
            errors.append('fim inválido. Use horário no formato HH:MM.')
        if start_time and end_time and end_time <= start_time:
            errors.append('fim deve ser posterior a inicio.')

        title = _validate_text('descricao', _row_value(row, header_map, 'descricao'), 300, errors, required=True)
        expected_result = _validate_text('resultado_esperado', _row_value(row, header_map, 'resultado_esperado'), 300, errors, required=True)
        evidence_required = _validate_text('evidencia_exigida', _row_value(row, header_map, 'evidencia_exigida'), 300, errors)
        proof_location = _validate_text('onde_comprovar', _row_value(row, header_map, 'onde_comprovar'), 300, errors)
        notes = _validate_text('observacoes', _row_value(row, header_map, 'observacoes'), 2000, errors)

        frequency = _parse_frequency(_row_value(row, header_map, 'recorrencia'))
        if not frequency:
            errors.append('recorrencia inválida. Use Diária, Semanal, Mensal ou Anual.')

        active = _parse_active(_row_value(row, header_map, 'ativo'))
        if active is None:
            errors.append('ativo inválido. Use Sim/Não, Ativo/Inativo ou True/False.')

        if position and day_of_week is not None and start_time and end_time and title:
            matches = TaskTemplate.objects.filter(
                position=position,
                day_of_week=day_of_week,
                start_time=start_time,
                end_time=end_time,
                title__iexact=title,
            )
            if matches.count() > 1:
                errors.append('Existe mais de uma atividade com mesmo tipo/cargo, dia, horário e descrição. Ajuste manualmente antes de importar.')

            import_key = (position.pk, day_of_week, start_time, end_time, title.casefold())
            if import_key in seen_keys:
                errors.append(f'Linha duplicada no XLSX. A mesma atividade já aparece na linha {seen_keys[import_key]}.')
            else:
                seen_keys[import_key] = excel_row

        if errors:
            errors_by_row.append({'row': excel_row, 'errors': errors})
            continue

        parsed_rows.append(ActivityImportRow(
            excel_row=excel_row,
            position=position,
            day_of_week=day_of_week,
            start_time=start_time,
            end_time=end_time,
            title=title,
            expected_result=expected_result,
            frequency=frequency,
            evidence_required=evidence_required,
            proof_location=proof_location,
            notes=notes,
            active=active,
        ))

    if not parsed_rows and not errors_by_row:
        errors_by_row.append({'row': 'Arquivo', 'errors': ['Nenhuma linha de atividade encontrada para importação.']})

    return parsed_rows, errors_by_row


def import_activity_rows(rows, actor, filename=''):
    created = 0
    updated = 0

    with transaction.atomic():
        for row in rows:
            matches = TaskTemplate.objects.filter(
                position=row.position,
                day_of_week=row.day_of_week,
                start_time=row.start_time,
                end_time=row.end_time,
                title__iexact=row.title,
            )
            template = matches.first()
            if template:
                template.title = row.title
                template.expected_result = row.expected_result
                template.frequency = row.frequency
                template.evidence_required = row.evidence_required
                template.proof_location = row.proof_location
                template.notes = row.notes
                template.active = row.active
                template.source = 'Importação XLSX'
                template.save(update_fields=[
                    'title', 'expected_result', 'frequency', 'evidence_required',
                    'proof_location', 'notes', 'active', 'source',
                ])
                updated += 1
            else:
                max_order = TaskTemplate.objects.filter(
                    position=row.position,
                    day_of_week=row.day_of_week,
                ).aggregate(max_order=Max('order'))['max_order'] or 0
                template = TaskTemplate.objects.create(
                    position=row.position,
                    day_of_week=row.day_of_week,
                    start_time=row.start_time,
                    end_time=row.end_time,
                    title=row.title,
                    expected_result=row.expected_result,
                    frequency=row.frequency,
                    evidence_required=row.evidence_required,
                    proof_location=row.proof_location,
                    notes=row.notes,
                    active=row.active,
                    source='Importação XLSX',
                    order=max_order + 1,
                )
                created += 1

        ActivityLog.objects.create(
            actor=actor,
            action='Importação XLSX de atividades',
            object_type='TaskTemplate',
            details=(
                f'Arquivo: {filename or "sem nome"}; linhas processadas: {len(rows)}; '
                f'criadas: {created}; atualizadas: {updated}'
            ),
        )

    return ActivityImportResult(created=created, updated=updated, processed=len(rows))


def build_activity_import_template():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Atividades'
    worksheet.append(ACTIVITY_IMPORT_HEADERS)
    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = f'A1:{get_column_letter(len(ACTIVITY_IMPORT_HEADERS))}1'

    header_fill = PatternFill('solid', fgColor='1F6FEB')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    widths = {
        'A': 24,
        'B': 28,
        'C': 18,
        'D': 12,
        'E': 12,
        'F': 42,
        'G': 42,
        'H': 16,
        'I': 34,
        'J': 34,
        'K': 48,
        'L': 12,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    weekday_validation = DataValidation(
        type='list',
        formula1='"Segunda-feira,Terça-feira,Quarta-feira,Quinta-feira,Sexta-feira,Sábado"',
        allow_blank=False,
    )
    frequency_validation = DataValidation(type='list', formula1='"Diária,Semanal,Mensal,Anual"', allow_blank=False)
    active_validation = DataValidation(type='list', formula1='"Sim,Não"', allow_blank=True)
    worksheet.add_data_validation(weekday_validation)
    worksheet.add_data_validation(frequency_validation)
    worksheet.add_data_validation(active_validation)
    weekday_validation.add('C2:C501')
    frequency_validation.add('H2:H501')
    active_validation.add('L2:L501')

    instructions = workbook.create_sheet('Instrucoes')
    instructions['A1'] = 'Modelo de importação de atividades'
    instructions['A1'].font = Font(bold=True, size=14)
    rows = [
        'Preencha a aba Atividades a partir da linha 2.',
        'Informe tipo_usuario_codigo ou tipo_usuario_nome. Se preencher ambos, eles devem apontar para o mesmo tipo/cargo.',
        'Campos obrigatórios: tipo/cargo, dia_semana, inicio, fim, descricao, resultado_esperado e recorrencia.',
        'Valores aceitos em dia_semana: Segunda-feira, Terça-feira, Quarta-feira, Quinta-feira, Sexta-feira, Sábado.',
        'Valores aceitos em recorrencia: Diária, Semanal, Mensal, Anual.',
        'Use horários no formato HH:MM.',
        'Ativo aceita Sim/Não, Ativo/Inativo ou True/False. Em branco será tratado como Sim.',
        'A importação cria ou atualiza pelo conjunto: tipo/cargo, dia, início, fim e descrição. Linhas ausentes no XLSX não excluem atividades.',
    ]
    for index, text in enumerate(rows, start=3):
        instructions[f'A{index}'] = text
    instructions.column_dimensions['A'].width = 120

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
