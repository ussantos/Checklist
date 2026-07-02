import re
from dataclasses import dataclass, field
from decimal import Decimal

from django.utils.text import slugify
from openpyxl import load_workbook

from .grafana_sync import clear_metric_dashboards, sync_metric_area
from .models import MetricRecord, MetricType, Position


@dataclass
class MetricImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)
    synced_areas: list[str] = field(default_factory=list)
    cleared_dashboards: list[str] = field(default_factory=list)
    deleted_metrics: int = 0
    deleted_records: int = 0


def _decimal_from_text(value):
    if value is None:
        return None
    text = str(value)
    match = re.search(r'[-+]?\d+(?:[,.]\d+)?', text)
    if not match:
        return None
    return Decimal(match.group(0).replace(',', '.'))


def _numbers_from_text(value):
    if value is None:
        return []
    return [Decimal(item.replace(',', '.')) for item in re.findall(r'[-+]?\d+(?:[,.]\d+)?', str(value))]


def infer_metric_metadata(name, target_text):
    text = f'{name or ""} {target_text or ""}'.lower()
    numbers = _numbers_from_text(target_text)
    metadata = {
        'frequency': MetricType.FREQ_MONTHLY,
        'unit': 'un',
        'target_direction': MetricType.TARGET_AT_LEAST,
        'target_min': None,
        'target_max': None,
        'monthly_target': Decimal('0'),
    }

    if 'semana' in text:
        metadata['frequency'] = MetricType.FREQ_WEEKLY
    if '%' in text or 'retenção' in text or 'presença' in text or 'promotores' in text:
        metadata['unit'] = '%'
    elif 'minuto' in text:
        metadata['unit'] = 'min'
    elif 'lead' in text:
        metadata['unit'] = 'leads'

    if 'até' in text:
        metadata['target_direction'] = MetricType.TARGET_AT_MOST
    elif ' a ' in text and len(numbers) >= 2:
        metadata['target_direction'] = MetricType.TARGET_RANGE
        metadata['target_min'] = numbers[0]
        metadata['target_max'] = numbers[1]
    elif text.strip().startswith('+'):
        metadata['target_direction'] = MetricType.TARGET_INCREASE
    else:
        metadata['target_direction'] = MetricType.TARGET_AT_LEAST

    if numbers:
        metadata['monthly_target'] = numbers[0]
        if metadata['target_max'] is not None:
            metadata['monthly_target'] = metadata['target_max']
        if metadata['target_min'] is None and metadata['target_direction'] == MetricType.TARGET_AT_LEAST:
            metadata['target_min'] = numbers[0]
        if metadata['target_max'] is None and metadata['target_direction'] == MetricType.TARGET_AT_MOST:
            metadata['target_max'] = numbers[0]

    return metadata


def infer_formula(name):
    normalized = (name or '').lower()
    if 'leads por semana' in normalized:
        return 'Contagem de oportunidades criadas na semana.'
    if 'leads que viram aula experimental' in normalized:
        return 'Aulas experimentais agendadas / leads recebidos x 100.'
    if 'comparecimento' in normalized:
        return 'Aulas experimentais com presença / aulas experimentais agendadas x 100.'
    if 'visita' in normalized and 'matr' in normalized:
        return 'Oportunidades ganhas / visitas ou aulas experimentais realizadas x 100.'
    if 'avaliação google' in normalized:
        return 'Solicitações de avaliação Google realizadas / oportunidades elegíveis x 100.'
    if 'conversão adicional' in normalized:
        return 'Conversões recuperadas por follow-up / leads indecisos x 100.'
    if 'tempo médio' in normalized:
        return 'Tempo entre decisão de matrícula e conclusão operacional do cadastro/pagamento.'
    if 'renovação' in normalized:
        return 'Renovações ou recompras / alunos elegíveis para renovação x 100.'
    if 'frequência' in normalized:
        return 'Aulas com presença / aulas previstas x 100.'
    if normalized == 'nps':
        return 'Promotores / respostas válidas da pesquisa NPS x 100.'
    return 'Cálculo definido pelo administrador no Checklist.'


def infer_data_source(name):
    normalized = (name or '').lower()
    if 'frequência' in normalized:
        return 'Checklist / Sponte - Agenda de aulas'
    if 'nps' in normalized or 'avaliação' in normalized:
        return 'Checklist - Pesquisas e registros pós-venda'
    if 'tempo médio' in normalized:
        return 'Checklist - Oportunidades e matrícula'
    return 'Checklist - Oportunidades comerciais'


def _unique_code(position, area, name, current_pk=None):
    base = slugify(f'{position.code}-{area}-{name}')[:70] or 'indicador'
    code = base
    suffix = 2
    qs = MetricType.objects.all()
    if current_pk:
        qs = qs.exclude(pk=current_pk)
    while qs.filter(code=code).exists():
        suffix_text = f'-{suffix}'
        code = f'{base[:80 - len(suffix_text)]}{suffix_text}'
        suffix += 1
    return code


def _row_dict(headers, row):
    return {str(header or '').strip(): value for header, value in zip(headers, row)}


def import_metrics_xlsx(path, *, sync_grafana=True, replace_existing=False):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(value or '').strip() for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    required = {'Etapa', 'Descrição', 'Indicador Principal', 'Meta'}
    missing = required - set(headers)
    if missing:
        raise ValueError(f'Planilha sem colunas obrigatórias: {", ".join(sorted(missing))}')

    position, _ = Position.objects.get_or_create(
        code='atendente-comercial',
        defaults={'name': 'Atendente Comercial', 'active': True},
    )
    result = MetricImportResult()
    touched_areas = set()

    if replace_existing:
        if sync_grafana:
            result.cleared_dashboards = clear_metric_dashboards()
        result.deleted_records, _ = MetricRecord.objects.all().delete()
        result.deleted_metrics, _ = MetricType.objects.all().delete()

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = _row_dict(headers, row)
        name = str(data.get('Indicador Principal') or '').strip()
        stage = str(data.get('Etapa') or '').strip()
        if not name or not stage:
            result.skipped += 1
            result.errors.append(f'Linha {row_number}: etapa ou indicador vazio.')
            continue

        target_text = str(data.get('Meta') or '').strip()
        metadata = infer_metric_metadata(name, target_text)
        area = stage[:70]
        code_seed = slugify(f'{position.code}-{area}-{name}')[:70] or 'indicador'
        metric = MetricType.objects.filter(position=position, area=area, name=name[:140]).first()
        if metric is None:
            metric = MetricType.objects.filter(code=code_seed).first()
        created = metric is None
        if metric is None:
            metric = MetricType(code=code_seed)
        metric.name = name[:140]
        metric.area = area
        metric.frequency = metadata['frequency']
        metric.position = position
        metric.description = str(data.get('Descrição') or '').strip()
        metric.formula = infer_formula(name)
        metric.data_source = infer_data_source(name)
        metric.monthly_target = metadata['monthly_target']
        metric.target_text = target_text[:180]
        metric.target_min = metadata['target_min']
        metric.target_max = metadata['target_max']
        metric.target_direction = metadata['target_direction']
        metric.unit = metadata['unit']
        metric.deliverables = str(data.get('Entregáveis') or '').strip()
        metric.attention_points = str(data.get('Pontos de Atenção') or '').strip()
        metric.visualization_type = MetricType.VIZ_AUTO
        metric.active = True
        metric.code = _unique_code(position, area, name, current_pk=metric.pk) if created else metric.code
        metric.save()
        touched_areas.add(metric.area)
        if created:
            result.created += 1
        else:
            result.updated += 1

    if sync_grafana:
        for area in sorted(touched_areas):
            sync_metric_area(area)
            result.synced_areas.append(area)

    return result
