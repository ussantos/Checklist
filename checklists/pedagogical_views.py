from datetime import date, timedelta
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.db.models import Q
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils.text import slugify
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .audit import changed_values, log_activity, snapshot_instance
from .forms import (
    CourseForm, LessonFeedbackForm, LessonForm, PedagogicalStudentForm, RoomForm, SchoolHolidayForm,
    SponteLessonReportXMLUploadForm, TimeSlotForm,
)
from .models import (
    CommercialOpportunity, Course, Lesson, LessonFeedback, PedagogicalReportTask,
    PedagogicalStudent, Room, SchoolHoliday, SponteSyncJob, TimeSlot,
)
from .services import (
    create_pedagogical_report_task_for_feedback,
    create_qualified_sale_opportunity_for_lesson_feedback, get_user_position, is_admin_user,
)
from .sponte import SponteClientError, default_sponte_schedule_window, reconcile_sponte_free_class_report_xml
from .sponte_jobs import enqueue_sponte_sync_job


COURSE_AUDIT_FIELDS = [
    'name', 'description', 'value', 'kit_quantity', 'max_students_per_slot', 'requires_module_count',
    'active', 'source', 'external_id', 'synced_at',
]
ROOM_AUDIT_FIELDS = ['name', 'capacity', 'active']
TIME_SLOT_AUDIT_FIELDS = ['weekday', 'start_time', 'end_time', 'active']
HOLIDAY_AUDIT_FIELDS = ['description', 'kind', 'start_date', 'end_date', 'active']
STUDENT_AUDIT_FIELDS = ['name', 'enrollment_number', 'responsible_name', 'whatsapp', 'status', 'source', 'external_id']
LESSON_AUDIT_FIELDS = [
    'lesson_type', 'trial_kind', 'commercial_opportunity', 'student', 'student_name_snapshot', 'responsible_name_snapshot',
    'whatsapp_snapshot', 'course', 'room', 'date', 'start_time', 'end_time',
    'status', 'notes', 'source', 'external_id', 'synced_at', 'created_by',
]
LESSON_FEEDBACK_AUDIT_FIELDS = [
    'lesson', 'punctuality_score', 'assembly_comment', 'assembly_score',
    'has_programming', 'programming_comment', 'programming_score',
    'participation_comment', 'participation_score', 'behavior_comment',
    'behavior_score', 'general_comment', 'general_score', 'created_by', 'updated_by',
]
PEDAGOGICAL_REPORT_TASK_AUDIT_FIELDS = [
    'feedback', 'student', 'course', 'module_number', 'lesson_number',
    'due_date', 'completed', 'completed_at', 'completed_by', 'created_by',
]
INSTRUCTOR_POSITION_CODE = 'instrutor-aula-livre'
COMMERCIAL_POSITION_CODE = 'atendente-comercial'


def _admin_check(user):
    return user.is_authenticated and is_admin_user(user)


def _feedback_access_check(user):
    if not user.is_authenticated:
        return False
    if is_admin_user(user):
        return True
    position = get_user_position(user)
    return bool(position and position.code == INSTRUCTOR_POSITION_CODE)


def _instructor_check(user):
    if not user.is_authenticated:
        return False
    position = get_user_position(user)
    return bool(position and position.code == INSTRUCTOR_POSITION_CODE and not is_admin_user(user))


def _sponte_import_check(user):
    if not user.is_authenticated:
        return False
    if is_admin_user(user):
        return True
    position = get_user_position(user)
    return bool(position and position.code == INSTRUCTOR_POSITION_CODE)


def _commercial_operator_check(user):
    if not user.is_authenticated:
        return False
    position = get_user_position(user)
    return bool(position and position.code == COMMERCIAL_POSITION_CODE and not is_admin_user(user))


def _parse_date(value):
    if not value:
        return timezone.localdate()
    try:
        return date.fromisoformat(value)
    except ValueError:
        return timezone.localdate()


def _parse_optional_date(value):
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def _agenda_period(value):
    return value if value in {'hoje', 'semana', 'mes'} else 'hoje'


def _agenda_period_bounds(period, target_date):
    period = _agenda_period(period)
    if period == 'semana':
        start_date = target_date - timedelta(days=target_date.weekday())
        return start_date, start_date + timedelta(days=6)
    if period == 'mes':
        start_date = target_date.replace(day=1)
        if target_date.month == 12:
            next_month = target_date.replace(year=target_date.year + 1, month=1, day=1)
        else:
            next_month = target_date.replace(month=target_date.month + 1, day=1)
        return start_date, next_month - timedelta(days=1)
    return target_date, target_date


def _active_filter(queryset, status, *, active_label='ativos', inactive_label='inativos', all_label='todos'):
    if status == inactive_label:
        return queryset.filter(active=False), status
    if status == all_label:
        return queryset, status
    return queryset.filter(active=True), active_label


def _student_status_filter(queryset, status):
    if status == 'inativos':
        return queryset.filter(status=PedagogicalStudent.STATUS_INACTIVE), status
    if status == 'todos':
        return queryset, status
    return queryset.filter(status=PedagogicalStudent.STATUS_ACTIVE), 'ativos'


def _log_change(*, request, obj, action, audit_fields, previous_values_full=None, details=''):
    previous_values = {}
    new_values = snapshot_instance(obj, audit_fields)
    if previous_values_full is not None:
        previous_values, new_values = changed_values(previous_values_full, new_values)
    log_activity(
        actor=request.user,
        obj=obj,
        action=action,
        object_label=str(obj),
        details=details,
        previous_values=previous_values,
        new_values=new_values,
    )


def _posted_next_url(request, fallback):
    target = request.POST.get('next') or fallback
    if url_has_allowed_host_and_scheme(target, allowed_hosts={request.get_host()}, require_https=request.is_secure()):
        return target
    return fallback


def _notify_sponte_job(request, job, created):
    period = f' Período: {job.period_label}.' if job.period_label else ''
    if created:
        messages.info(
            request,
            f'Sincronização Sponte de {job.get_kind_display().lower()} iniciada em segundo plano.{period} '
            'Você pode continuar usando o sistema.',
        )
    else:
        messages.warning(
            request,
            f'Já existe uma sincronização Sponte de {job.get_kind_display().lower()} em andamento.{period}',
        )


def _feedback_lessons_queryset():
    return (
        Lesson.objects
        .filter(
            lesson_type=Lesson.TYPE_REGULAR,
            status=Lesson.STATUS_DONE,
            student__status=PedagogicalStudent.STATUS_ACTIVE,
        )
        .select_related('student', 'course', 'room', 'feedback', 'feedback__created_by', 'feedback__updated_by')
        .order_by('-date', '-start_time', 'student_name_snapshot')
    )


def _feedback_due_lessons_queryset():
    today = timezone.localdate()
    current_time = timezone.localtime().time()
    return _feedback_lessons_queryset().filter(
        Q(date__lt=today) | Q(date=today, end_time__lte=current_time)
    )


def _attach_lesson_display_state(lessons):
    for lesson in lessons:
        lesson.assistant_warning = lesson.assistant_warning_message()
        try:
            lesson.feedback_record = lesson.feedback
        except LessonFeedback.DoesNotExist:
            lesson.feedback_record = None
    return lessons


def _instructor_feedback_queryset(user):
    return (
        LessonFeedback.objects
        .filter(created_by=user)
        .select_related('lesson', 'lesson__student', 'lesson__course', 'lesson__room', 'created_by', 'updated_by')
        .order_by('-lesson__date', '-lesson__start_time', 'lesson__student_name_snapshot')
    )


def _instructor_feedback_students(user):
    return (
        PedagogicalStudent.objects
        .filter(
            lessons__feedback__created_by=user,
            status=PedagogicalStudent.STATUS_ACTIVE,
        )
        .distinct()
        .order_by('name')
    )


def _instructor_report_tasks_queryset(user):
    return (
        PedagogicalReportTask.objects
        .filter(
            Q(created_by=user) | Q(feedback__created_by=user) | Q(feedback__updated_by=user),
            student__status=PedagogicalStudent.STATUS_ACTIVE,
        )
        .select_related(
            'feedback', 'feedback__lesson', 'student', 'course', 'completed_by', 'created_by',
        )
        .distinct()
        .order_by('completed', 'due_date', 'student__name', 'module_number', 'lesson_number')
    )


def _feedback_export_workbook(feedbacks):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Feedbacks'

    headers = [
        'Data', 'Horário', 'Aluno', 'Responsável', 'WhatsApp', 'Curso', 'Sala',
        'Módulo', 'Aula', 'Pontualidade', 'Montagem', 'Montagem nota',
        'Tem programação?', 'Programação', 'Programação nota',
        'Interesse/Participação', 'Interesse/Participação nota',
        'Comportamento', 'Comportamento nota', 'Comentário geral',
        'Comentário geral nota', 'Criado em', 'Atualizado em',
    ]
    sheet.append(headers)
    header_fill = PatternFill('solid', fgColor='1F4E78')
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for feedback in feedbacks:
        lesson = feedback.lesson
        created_at = timezone.localtime(feedback.created_at).replace(tzinfo=None) if feedback.created_at else None
        updated_at = timezone.localtime(feedback.updated_at).replace(tzinfo=None) if feedback.updated_at else None
        sheet.append([
            lesson.date,
            f'{lesson.start_time:%H:%M} - {lesson.end_time:%H:%M}',
            lesson.student_name_snapshot,
            lesson.responsible_name_snapshot,
            lesson.whatsapp_snapshot,
            lesson.course.name if lesson.course_id else '',
            lesson.room.name if lesson.room_id else '',
            feedback.module_number,
            feedback.lesson_number,
            dict(LessonFeedback.PUNCTUALITY_CHOICES).get(feedback.punctuality_score, feedback.punctuality_score),
            feedback.assembly_comment,
            feedback.assembly_score,
            'Sim' if feedback.has_programming else 'Não',
            feedback.programming_comment,
            feedback.programming_score,
            feedback.participation_comment,
            feedback.participation_score,
            feedback.behavior_comment,
            feedback.behavior_score,
            feedback.general_comment,
            float(feedback.general_score),
            created_at,
            updated_at,
        ])

    for row in sheet.iter_rows(min_row=2):
        row[0].number_format = 'dd/mm/yyyy'
        row[20].number_format = '0.00'
        if row[21].value:
            row[21].number_format = 'dd/mm/yyyy hh:mm'
        if row[22].value:
            row[22].number_format = 'dd/mm/yyyy hh:mm'
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    widths = {
        'A': 12, 'B': 16, 'C': 28, 'D': 28, 'E': 18, 'F': 20, 'G': 18,
        'H': 10, 'I': 10, 'J': 14, 'K': 44, 'L': 14, 'M': 16, 'N': 44,
        'O': 16, 'P': 44, 'Q': 24, 'R': 44, 'S': 18, 'T': 52,
        'U': 18, 'V': 18, 'W': 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = 'A2'

    return workbook


@user_passes_test(_admin_check)
def courses_list(request):
    status = request.GET.get('status', 'ativos')
    courses, status = _active_filter(Course.objects.all().order_by('-active', 'name'), status)
    return render(request, 'checklists/courses_list.html', {
        'courses': courses,
        'status': status,
        'is_admin': True,
    })


@require_POST
@user_passes_test(_admin_check)
def courses_sync_sponte(request):
    job, created = enqueue_sponte_sync_job(
        kind=SponteSyncJob.KIND_COURSES,
        requested_by=request.user,
    )
    _notify_sponte_job(request, job, created)
    return redirect('pedagogical_courses')


@user_passes_test(_admin_check)
def course_create(request):
    if request.method == 'POST':
        form = CourseForm(request.POST)
        if form.is_valid():
            course = form.save()
            _log_change(
                request=request,
                obj=course,
                action='Curso criado',
                audit_fields=COURSE_AUDIT_FIELDS,
                details=(
                    f'Curso: {course.name}; valor: {course.value}; kits: {course.kit_quantity}; '
                    f'módulos: {course.requires_module_count}; ativo: {course.active}'
                ),
            )
            messages.success(request, 'Curso criado.')
            return redirect('pedagogical_courses')
    else:
        form = CourseForm()

    return render(request, 'checklists/course_form.html', {
        'form': form,
        'title': 'Inserir curso',
        'submit_label': 'Inserir curso',
        'is_admin': True,
    })


@user_passes_test(_admin_check)
def course_edit(request, course_id):
    course = get_object_or_404(Course, pk=course_id)
    previous_active = course.active
    previous_values_full = snapshot_instance(course, COURSE_AUDIT_FIELDS)
    if request.method == 'POST':
        form = CourseForm(request.POST, instance=course)
        if form.is_valid():
            course = form.save()
            action = 'Curso atualizado'
            if previous_active and not course.active:
                action = 'Curso desativado'
            elif not previous_active and course.active:
                action = 'Curso ativado'
            _log_change(
                request=request,
                obj=course,
                action=action,
                audit_fields=COURSE_AUDIT_FIELDS,
                previous_values_full=previous_values_full,
                details=(
                    f'Curso: {course.name}; valor: {course.value}; kits: {course.kit_quantity}; '
                    f'módulos: {course.requires_module_count}; ativo: {course.active}'
                ),
            )
            messages.success(request, 'Curso atualizado.')
            return redirect('pedagogical_courses')
    else:
        form = CourseForm(instance=course)

    return render(request, 'checklists/course_form.html', {
        'form': form,
        'title': f'Editar curso - {course.name}',
        'submit_label': 'Salvar alterações',
        'course': course,
        'is_admin': True,
    })


@require_POST
@user_passes_test(_admin_check)
def course_toggle(request, course_id):
    course = get_object_or_404(Course, pk=course_id)
    previous_values_full = snapshot_instance(course, COURSE_AUDIT_FIELDS)
    course.active = not course.active
    course.save(update_fields=['active', 'updated_at'])
    _log_change(
        request=request,
        obj=course,
        action='Curso ativado' if course.active else 'Curso desativado',
        audit_fields=COURSE_AUDIT_FIELDS,
        previous_values_full=previous_values_full,
        details=(
            f'Curso: {course.name}; valor: {course.value}; kits: {course.kit_quantity}; '
            f'módulos: {course.requires_module_count}; ativo: {course.active}'
        ),
    )
    messages.success(request, f'Curso {"ativado" if course.active else "desativado"}.')
    return redirect('pedagogical_courses')


@user_passes_test(_admin_check)
def rooms_list(request):
    status = request.GET.get('status', 'ativas')
    rooms, status = _active_filter(
        Room.objects.all().order_by('-active', 'name'),
        status,
        active_label='ativas',
        inactive_label='inativas',
        all_label='todas',
    )
    return render(request, 'checklists/rooms_list.html', {
        'rooms': rooms,
        'status': status,
        'is_admin': True,
    })


@user_passes_test(_admin_check)
def room_create(request):
    if request.method == 'POST':
        form = RoomForm(request.POST)
        if form.is_valid():
            room = form.save()
            _log_change(
                request=request,
                obj=room,
                action='Sala criada',
                audit_fields=ROOM_AUDIT_FIELDS,
                details=f'Sala: {room.name}; capacidade: {room.capacity}; ativa: {room.active}',
            )
            messages.success(request, 'Sala criada.')
            return redirect('pedagogical_rooms')
    else:
        form = RoomForm()
    return render(request, 'checklists/room_form.html', {'form': form, 'title': 'Inserir sala', 'submit_label': 'Inserir sala', 'is_admin': True})


@user_passes_test(_admin_check)
def room_edit(request, room_id):
    room = get_object_or_404(Room, pk=room_id)
    previous_active = room.active
    previous_values_full = snapshot_instance(room, ROOM_AUDIT_FIELDS)
    if request.method == 'POST':
        form = RoomForm(request.POST, instance=room)
        if form.is_valid():
            room = form.save()
            action = 'Sala atualizada'
            if previous_active and not room.active:
                action = 'Sala desativada'
            elif not previous_active and room.active:
                action = 'Sala ativada'
            _log_change(
                request=request,
                obj=room,
                action=action,
                audit_fields=ROOM_AUDIT_FIELDS,
                previous_values_full=previous_values_full,
                details=f'Sala: {room.name}; capacidade: {room.capacity}; ativa: {room.active}',
            )
            messages.success(request, 'Sala atualizada.')
            return redirect('pedagogical_rooms')
    else:
        form = RoomForm(instance=room)
    return render(request, 'checklists/room_form.html', {'form': form, 'title': f'Editar sala - {room.name}', 'submit_label': 'Salvar alterações', 'room': room, 'is_admin': True})


@require_POST
@user_passes_test(_admin_check)
def room_toggle(request, room_id):
    room = get_object_or_404(Room, pk=room_id)
    previous_values_full = snapshot_instance(room, ROOM_AUDIT_FIELDS)
    room.active = not room.active
    room.save(update_fields=['active', 'updated_at'])
    _log_change(
        request=request,
        obj=room,
        action='Sala ativada' if room.active else 'Sala desativada',
        audit_fields=ROOM_AUDIT_FIELDS,
        previous_values_full=previous_values_full,
        details=f'Sala: {room.name}; capacidade: {room.capacity}; ativa: {room.active}',
    )
    messages.success(request, f'Sala {"ativada" if room.active else "desativada"}.')
    return redirect('pedagogical_rooms')


@user_passes_test(_admin_check)
def time_slots_list(request):
    status = request.GET.get('status', 'ativos')
    slots, status = _active_filter(TimeSlot.objects.all().order_by('weekday', 'start_time'), status)
    return render(request, 'checklists/time_slots_list.html', {
        'slots': slots,
        'status': status,
        'is_admin': True,
    })


@user_passes_test(_admin_check)
def time_slot_create(request):
    if request.method == 'POST':
        form = TimeSlotForm(request.POST)
        if form.is_valid():
            slot = form.save()
            _log_change(request=request, obj=slot, action='Horário criado', audit_fields=TIME_SLOT_AUDIT_FIELDS, details=f'Horário: {slot}; ativo: {slot.active}')
            messages.success(request, 'Horário criado.')
            return redirect('pedagogical_time_slots')
    else:
        form = TimeSlotForm()
    return render(request, 'checklists/time_slot_form.html', {'form': form, 'title': 'Inserir horário', 'submit_label': 'Inserir horário', 'is_admin': True})


@user_passes_test(_admin_check)
def time_slot_edit(request, slot_id):
    slot = get_object_or_404(TimeSlot, pk=slot_id)
    previous_active = slot.active
    previous_values_full = snapshot_instance(slot, TIME_SLOT_AUDIT_FIELDS)
    if request.method == 'POST':
        form = TimeSlotForm(request.POST, instance=slot)
        if form.is_valid():
            slot = form.save()
            action = 'Horário atualizado'
            if previous_active and not slot.active:
                action = 'Horário desativado'
            elif not previous_active and slot.active:
                action = 'Horário ativado'
            _log_change(request=request, obj=slot, action=action, audit_fields=TIME_SLOT_AUDIT_FIELDS, previous_values_full=previous_values_full, details=f'Horário: {slot}; ativo: {slot.active}')
            messages.success(request, 'Horário atualizado.')
            return redirect('pedagogical_time_slots')
    else:
        form = TimeSlotForm(instance=slot)
    return render(request, 'checklists/time_slot_form.html', {'form': form, 'title': f'Editar horário - {slot}', 'submit_label': 'Salvar alterações', 'slot': slot, 'is_admin': True})


@require_POST
@user_passes_test(_admin_check)
def time_slot_toggle(request, slot_id):
    slot = get_object_or_404(TimeSlot, pk=slot_id)
    previous_values_full = snapshot_instance(slot, TIME_SLOT_AUDIT_FIELDS)
    slot.active = not slot.active
    slot.save(update_fields=['active', 'updated_at'])
    _log_change(request=request, obj=slot, action='Horário ativado' if slot.active else 'Horário desativado', audit_fields=TIME_SLOT_AUDIT_FIELDS, previous_values_full=previous_values_full, details=f'Horário: {slot}; ativo: {slot.active}')
    messages.success(request, f'Horário {"ativado" if slot.active else "desativado"}.')
    return redirect('pedagogical_time_slots')


@user_passes_test(_admin_check)
def holidays_list(request):
    status = request.GET.get('status', 'ativos')
    holidays, status = _active_filter(SchoolHoliday.objects.all().order_by('start_date'), status)
    return render(request, 'checklists/holidays_list.html', {
        'holidays': holidays,
        'status': status,
        'is_admin': True,
    })


@user_passes_test(_admin_check)
def holiday_create(request):
    if request.method == 'POST':
        form = SchoolHolidayForm(request.POST)
        if form.is_valid():
            holiday = form.save()
            _log_change(request=request, obj=holiday, action='Feriado/Recesso criado', audit_fields=HOLIDAY_AUDIT_FIELDS, details=f'Feriado/Recesso: {holiday}; ativo: {holiday.active}')
            messages.success(request, 'Feriado/Recesso criado.')
            return redirect('pedagogical_holidays')
    else:
        form = SchoolHolidayForm()
    return render(request, 'checklists/holiday_form.html', {'form': form, 'title': 'Inserir feriado/recesso', 'submit_label': 'Inserir', 'is_admin': True})


@user_passes_test(_admin_check)
def holiday_edit(request, holiday_id):
    holiday = get_object_or_404(SchoolHoliday, pk=holiday_id)
    previous_active = holiday.active
    previous_values_full = snapshot_instance(holiday, HOLIDAY_AUDIT_FIELDS)
    if request.method == 'POST':
        form = SchoolHolidayForm(request.POST, instance=holiday)
        if form.is_valid():
            holiday = form.save()
            action = 'Feriado/Recesso atualizado'
            if previous_active and not holiday.active:
                action = 'Feriado/Recesso desativado'
            elif not previous_active and holiday.active:
                action = 'Feriado/Recesso ativado'
            _log_change(request=request, obj=holiday, action=action, audit_fields=HOLIDAY_AUDIT_FIELDS, previous_values_full=previous_values_full, details=f'Feriado/Recesso: {holiday}; ativo: {holiday.active}')
            messages.success(request, 'Feriado/Recesso atualizado.')
            return redirect('pedagogical_holidays')
    else:
        form = SchoolHolidayForm(instance=holiday)
    return render(request, 'checklists/holiday_form.html', {'form': form, 'title': f'Editar feriado/recesso - {holiday.description}', 'submit_label': 'Salvar alterações', 'holiday': holiday, 'is_admin': True})


@require_POST
@user_passes_test(_admin_check)
def holiday_toggle(request, holiday_id):
    holiday = get_object_or_404(SchoolHoliday, pk=holiday_id)
    previous_values_full = snapshot_instance(holiday, HOLIDAY_AUDIT_FIELDS)
    holiday.active = not holiday.active
    holiday.save(update_fields=['active', 'updated_at'])
    _log_change(request=request, obj=holiday, action='Feriado/Recesso ativado' if holiday.active else 'Feriado/Recesso desativado', audit_fields=HOLIDAY_AUDIT_FIELDS, previous_values_full=previous_values_full, details=f'Feriado/Recesso: {holiday}; ativo: {holiday.active}')
    messages.success(request, f'Feriado/Recesso {"ativado" if holiday.active else "desativado"}.')
    return redirect('pedagogical_holidays')


@user_passes_test(_admin_check)
def students_list(request):
    status = request.GET.get('status', 'ativos')
    students, status = _student_status_filter(PedagogicalStudent.objects.all().order_by('name'), status)
    return render(request, 'checklists/students_list.html', {
        'students': students,
        'status': status,
        'is_admin': True,
    })


@require_POST
@user_passes_test(_admin_check)
def students_import_sponte(request):
    job, created = enqueue_sponte_sync_job(
        kind=SponteSyncJob.KIND_STUDENTS,
        requested_by=request.user,
    )
    _notify_sponte_job(request, job, created)
    return redirect('pedagogical_students')


@user_passes_test(_admin_check)
def student_create(request):
    if request.method == 'POST':
        form = PedagogicalStudentForm(request.POST)
        if form.is_valid():
            student = form.save()
            _log_change(request=request, obj=student, action='Aluno pedagógico criado', audit_fields=STUDENT_AUDIT_FIELDS, details=f'Aluno: {student.name}; status: {student.status}')
            messages.success(request, 'Aluno criado.')
            return redirect('pedagogical_students')
    else:
        form = PedagogicalStudentForm()
    return render(request, 'checklists/student_form.html', {'form': form, 'title': 'Inserir aluno', 'submit_label': 'Inserir aluno', 'is_admin': True})


@user_passes_test(_admin_check)
def student_edit(request, student_id):
    student = get_object_or_404(PedagogicalStudent, pk=student_id)
    previous_status = student.status
    previous_values_full = snapshot_instance(student, STUDENT_AUDIT_FIELDS)
    if request.method == 'POST':
        form = PedagogicalStudentForm(request.POST, instance=student)
        if form.is_valid():
            student = form.save()
            action = 'Aluno pedagógico atualizado'
            if previous_status != student.status:
                action = 'Aluno pedagógico ativado' if student.active else 'Aluno pedagógico desativado'
            _log_change(request=request, obj=student, action=action, audit_fields=STUDENT_AUDIT_FIELDS, previous_values_full=previous_values_full, details=f'Aluno: {student.name}; status: {student.status}')
            messages.success(request, 'Aluno atualizado.')
            return redirect('pedagogical_students')
    else:
        form = PedagogicalStudentForm(instance=student)
    return render(request, 'checklists/student_form.html', {'form': form, 'title': f'Editar aluno - {student.name}', 'submit_label': 'Salvar alterações', 'student': student, 'is_admin': True})


@require_POST
@user_passes_test(_admin_check)
def student_toggle(request, student_id):
    student = get_object_or_404(PedagogicalStudent, pk=student_id)
    previous_values_full = snapshot_instance(student, STUDENT_AUDIT_FIELDS)
    student.status = PedagogicalStudent.STATUS_INACTIVE if student.active else PedagogicalStudent.STATUS_ACTIVE
    student.save(update_fields=['status', 'updated_at'])
    _log_change(request=request, obj=student, action='Aluno pedagógico ativado' if student.active else 'Aluno pedagógico desativado', audit_fields=STUDENT_AUDIT_FIELDS, previous_values_full=previous_values_full, details=f'Aluno: {student.name}; status: {student.status}')
    messages.success(request, f'Aluno {"ativado" if student.active else "desativado"}.')
    return redirect('pedagogical_students')


@user_passes_test(_feedback_access_check)
def lesson_feedbacks(request):
    today = timezone.localdate()
    legacy_date = _parse_optional_date(request.GET.get('data'))
    list_all = request.GET.get('listar') == 'todos'
    start_date = _parse_optional_date(request.GET.get('data_inicio')) or legacy_date or today
    end_date = _parse_optional_date(request.GET.get('data_fim')) or legacy_date or start_date
    if end_date < start_date:
        start_date, end_date = end_date, start_date
    feedback_status = request.GET.get('status', 'pendentes')
    if feedback_status not in {'pendentes', 'preenchidos', 'todos'}:
        feedback_status = 'pendentes'

    lessons = _feedback_lessons_queryset()
    if not list_all:
        lessons = lessons.filter(date__gte=start_date, date__lte=end_date)
    if feedback_status == 'pendentes':
        lessons = lessons.filter(feedback__isnull=True)
    elif feedback_status == 'preenchidos':
        lessons = lessons.filter(feedback__isnull=False)

    lessons = _attach_lesson_display_state(list(lessons))

    return render(request, 'checklists/lesson_feedbacks.html', {
        'lessons': lessons,
        'target_date': start_date,
        'start_date': start_date,
        'end_date': end_date,
        'start_date_value': '' if list_all else start_date.isoformat(),
        'end_date_value': '' if list_all else end_date.isoformat(),
        'list_all': list_all,
        'feedback_status': feedback_status,
        'feedback_students': _instructor_feedback_students(request.user) if not is_admin_user(request.user) else [],
        'is_admin': is_admin_user(request.user),
    })


@user_passes_test(_feedback_access_check)
def lesson_feedback_edit(request, lesson_id):
    lesson = get_object_or_404(
        _feedback_lessons_queryset(),
        pk=lesson_id,
    )
    if lesson.lesson_type != Lesson.TYPE_REGULAR:
        return HttpResponseForbidden('Feedback de aula é permitido apenas para aulas de aluno matriculado.')

    try:
        feedback = lesson.feedback
    except LessonFeedback.DoesNotExist:
        feedback = None
    previous_values_full = snapshot_instance(feedback, LESSON_FEEDBACK_AUDIT_FIELDS) if feedback else None
    if request.method == 'POST':
        form = LessonFeedbackForm(request.POST, instance=feedback)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.lesson = lesson
            if not feedback.pk:
                feedback.created_by = request.user
            feedback.updated_by = request.user
            feedback.save()
            action = 'Feedback de aula criado' if previous_values_full is None else 'Feedback de aula atualizado'
            _log_change(
                request=request,
                obj=feedback,
                action=action,
                audit_fields=LESSON_FEEDBACK_AUDIT_FIELDS,
                previous_values_full=previous_values_full,
                details=f'Feedback: {lesson}; nota geral: {feedback.general_score}',
            )
            report_task, report_created = create_pedagogical_report_task_for_feedback(feedback, actor=request.user)
            if report_created:
                messages.info(
                    request,
                    f'Relatório pedagógico criado com prazo em {report_task.due_date:%d/%m/%Y}.'
                )
            opportunity, created = create_qualified_sale_opportunity_for_lesson_feedback(feedback, actor=request.user)
            if created:
                messages.info(
                    request,
                    f'Oportunidade qualificada {opportunity.title} criada para follow-up amanhã.'
                )
            messages.success(request, 'Feedback de aula salvo.')
            return redirect(f'{reverse("pedagogical_lesson_feedbacks")}?data={lesson.date.isoformat()}')
    else:
        form = LessonFeedbackForm(instance=feedback)

    return render(request, 'checklists/lesson_feedback_form.html', {
        'form': form,
        'lesson': lesson,
        'feedback': feedback,
        'is_admin': is_admin_user(request.user),
    })


@user_passes_test(_instructor_check)
def instructor_feedback_export_xlsx(request):
    student = None
    feedbacks_queryset = _instructor_feedback_queryset(request.user)
    student_id = (request.GET.get('student') or '').strip()
    if student_id:
        student = get_object_or_404(_instructor_feedback_students(request.user), pk=student_id)
        feedbacks_queryset = feedbacks_queryset.filter(lesson__student=student)

    feedbacks = list(feedbacks_queryset)
    workbook = _feedback_export_workbook(feedbacks)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    if student:
        student_slug = slugify(student.name) or f'aluno-{student.pk}'
        filename = f'feedbacks_{student_slug}_{timezone.localdate():%Y%m%d}.xlsx'
    else:
        filename = f'feedbacks_instrutor_{timezone.localdate():%Y%m%d}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@require_POST
@user_passes_test(_instructor_check)
def instructor_report_task_complete(request, task_id):
    task = get_object_or_404(_instructor_report_tasks_queryset(request.user), pk=task_id)
    redirect_to = _posted_next_url(request, f'{reverse("instructor_dashboard")}#relatorios-pedagogicos')
    if task.completed:
        messages.info(request, 'Relatório pedagógico já estava marcado como concluído.')
        return redirect(redirect_to)

    previous_values_full = snapshot_instance(task, PEDAGOGICAL_REPORT_TASK_AUDIT_FIELDS)
    task.completed = True
    task.completed_at = timezone.now()
    task.completed_by = request.user
    task.save(update_fields=['completed', 'completed_at', 'completed_by', 'updated_at'])
    _log_change(
        request=request,
        obj=task,
        action='Relatório pedagógico concluído',
        audit_fields=PEDAGOGICAL_REPORT_TASK_AUDIT_FIELDS,
        previous_values_full=previous_values_full,
        details=f'Relatório pedagógico: {task.student.name}; módulo {task.module_number}; aula {task.lesson_number}.',
    )
    messages.success(request, 'Relatório pedagógico marcado como concluído.')
    return redirect(redirect_to)


def _lesson_queryset():
    return Lesson.objects.select_related(
        'student', 'commercial_opportunity', 'course', 'room', 'created_by',
    ).order_by('date', 'start_time', 'room__name')


def _required_instructor_lesson_queryset():
    return _lesson_queryset().filter(
        Q(lesson_type=Lesson.TYPE_TRIAL) |
        Q(student__status=PedagogicalStudent.STATUS_ACTIVE)
    )


def _apply_lesson_filters(request, lessons):
    course_id = request.GET.get('curso', '')
    room_id = request.GET.get('sala', '')
    lesson_type = request.GET.get('tipo', '')
    status = request.GET.get('status', '')
    if course_id.isdigit():
        lessons = lessons.filter(course_id=course_id)
    else:
        course_id = ''
    if room_id.isdigit():
        lessons = lessons.filter(room_id=room_id)
    else:
        room_id = ''
    if lesson_type in dict(Lesson.TYPE_CHOICES):
        lessons = lessons.filter(lesson_type=lesson_type)
    else:
        lesson_type = ''
    if status in dict(Lesson.STATUS_CHOICES):
        lessons = lessons.filter(status=status)
    else:
        status = ''
    return lessons, {
        'course_id': course_id,
        'room_id': room_id,
        'lesson_type': lesson_type,
        'status': status,
    }


def _lesson_calendar(lessons, period_start, period_end, period):
    lessons_by_date = {}
    for lesson in lessons:
        lessons_by_date.setdefault(lesson.date, []).append(lesson)

    if period == 'mes':
        calendar_start = period_start - timedelta(days=period_start.weekday())
        calendar_end = period_end + timedelta(days=6 - period_end.weekday())
    else:
        calendar_start = period_start
        calendar_end = period_end

    days = []
    current = calendar_start
    today = timezone.localdate()
    while current <= calendar_end:
        days.append({
            'date': current,
            'lessons': lessons_by_date.get(current, []),
            'in_period': period_start <= current <= period_end,
            'is_today': current == today,
        })
        current += timedelta(days=1)

    if period == 'hoje':
        return [days]
    return [days[index:index + 7] for index in range(0, len(days), 7)]


@user_passes_test(_instructor_check)
def instructor_dashboard(request):
    today = timezone.localdate()
    week_start, week_end = _agenda_period_bounds('semana', today)
    report_window_end = today + timedelta(days=15)
    current_time = timezone.localtime().time()

    today_lessons = _attach_lesson_display_state(list(
        _required_instructor_lesson_queryset()
        .filter(date=today)
        .order_by('start_time', 'room__name', 'student_name_snapshot')
    ))
    week_lessons = list(_required_instructor_lesson_queryset().filter(date__gte=week_start, date__lte=week_end))
    upcoming_lessons = _attach_lesson_display_state(list(
        _required_instructor_lesson_queryset()
        .filter(date__gte=today)
        .exclude(status=Lesson.STATUS_CANCELLED)
        .order_by('date', 'start_time', 'room__name')[:8]
    ))
    upcoming_trial_lessons = _attach_lesson_display_state(list(
        _required_instructor_lesson_queryset()
        .filter(
            lesson_type=Lesson.TYPE_TRIAL,
            date__gte=today,
            date__lte=week_end,
        )
        .exclude(status=Lesson.STATUS_CANCELLED)
        .order_by('date', 'start_time', 'room__name')[:6]
    ))
    pending_feedbacks_queryset = _feedback_due_lessons_queryset().filter(feedback__isnull=True)
    pending_feedbacks = _attach_lesson_display_state(list(
        pending_feedbacks_queryset.order_by('date', 'start_time', 'student_name_snapshot')[:8]
    ))

    next_lesson = next(
        (
            lesson for lesson in today_lessons
            if lesson.status != Lesson.STATUS_CANCELLED and lesson.end_time >= current_time
        ),
        None,
    )
    if next_lesson is None and upcoming_lessons:
        next_lesson = upcoming_lessons[0]

    pending_feedback_count = pending_feedbacks_queryset.count()
    overdue_feedback_count = pending_feedbacks_queryset.filter(date__lt=today).count()
    today_feedback_count = pending_feedbacks_queryset.filter(date=today).count()
    pending_report_tasks_queryset = (
        _instructor_report_tasks_queryset(request.user)
        .filter(completed=False, due_date__lte=report_window_end)
    )
    pending_report_tasks = list(pending_report_tasks_queryset[:8])
    report_tasks_count = pending_report_tasks_queryset.count()
    overdue_report_tasks_count = pending_report_tasks_queryset.filter(due_date__lt=today).count()
    today_report_tasks_count = pending_report_tasks_queryset.filter(due_date=today).count()

    return render(request, 'checklists/instructor_dashboard.html', {
        'today': today,
        'week_start': week_start,
        'week_end': week_end,
        'report_window_end': report_window_end,
        'today_lessons': today_lessons,
        'next_lesson': next_lesson,
        'pending_feedbacks': pending_feedbacks,
        'pending_report_tasks': pending_report_tasks,
        'upcoming_trial_lessons': upcoming_trial_lessons,
        'today_count': len(today_lessons),
        'week_count': len(week_lessons),
        'trial_week_count': sum(1 for lesson in week_lessons if lesson.lesson_type == Lesson.TYPE_TRIAL),
        'pending_feedback_count': pending_feedback_count,
        'today_feedback_count': today_feedback_count,
        'overdue_feedback_count': overdue_feedback_count,
        'report_tasks_count': report_tasks_count,
        'today_report_tasks_count': today_report_tasks_count,
        'overdue_report_tasks_count': overdue_report_tasks_count,
        'feedback_students': _instructor_feedback_students(request.user),
        'position': get_user_position(request.user),
        'is_admin': False,
    })


@user_passes_test(_instructor_check)
def instructor_agenda(request):
    target_date = _parse_date(request.GET.get('data'))
    period = _agenda_period(request.GET.get('periodo', 'semana'))
    period_start, period_end = _agenda_period_bounds(period, target_date)
    sponte_xml_form = SponteLessonReportXMLUploadForm(initial={
        'start_date': period_start,
        'end_date': period_end,
    })
    lessons = _required_instructor_lesson_queryset().filter(date__gte=period_start, date__lte=period_end)
    lessons, selected = _apply_lesson_filters(request, lessons)
    lessons = _attach_lesson_display_state(list(lessons))
    selected['period'] = period
    calendar_weeks = _lesson_calendar(lessons, period_start, period_end, period)
    return render(request, 'checklists/lesson_agenda.html', {
        'title': 'Agenda do Instrutor',
        'lessons': lessons,
        'calendar_weeks': calendar_weeks,
        'weekday_labels': ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom'],
        'target_date': target_date,
        'period_start': period_start,
        'period_end': period_end,
        'selected': selected,
        'courses': Course.objects.order_by('name'),
        'rooms': Room.objects.order_by('name'),
        'lesson_type_choices': Lesson.TYPE_CHOICES,
        'status_choices': Lesson.STATUS_CHOICES,
        'trial_only': False,
        'instructor_mode': True,
        'sponte_xml_form': sponte_xml_form,
        'sponte_xml_import_url': reverse('instructor_lessons_import_sponte_xml'),
        'is_admin': False,
    })


@require_POST
@user_passes_test(_sponte_import_check)
def instructor_import_sponte(request):
    redirect_to = _posted_next_url(request, reverse('instructor_dashboard'))
    today = timezone.localdate()
    start_date, end_date = default_sponte_schedule_window(today)
    job, created = enqueue_sponte_sync_job(
        kind=SponteSyncJob.KIND_INSTRUCTOR_SCHEDULE,
        requested_by=request.user,
        period_start=start_date,
        period_end=end_date,
    )
    _notify_sponte_job(request, job, created)
    return redirect(redirect_to)


@user_passes_test(_commercial_operator_check)
def commercial_lesson_agenda(request):
    target_date = _parse_date(request.GET.get('data'))
    period = _agenda_period(request.GET.get('periodo', 'semana'))
    period_start, period_end = _agenda_period_bounds(period, target_date)
    lessons = _lesson_queryset().filter(date__gte=period_start, date__lte=period_end)
    lessons, selected = _apply_lesson_filters(request, lessons)
    lessons = _attach_lesson_display_state(list(lessons))
    for lesson in lessons:
        lesson.can_open_opportunity = (
            lesson.commercial_opportunity_id
            and lesson.commercial_opportunity.owner_id == request.user.id
        )
    selected['period'] = period
    calendar_weeks = _lesson_calendar(lessons, period_start, period_end, period)
    return render(request, 'checklists/lesson_agenda.html', {
        'title': 'Agenda Comercial',
        'lessons': lessons,
        'calendar_weeks': calendar_weeks,
        'weekday_labels': ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom'],
        'target_date': target_date,
        'period_start': period_start,
        'period_end': period_end,
        'selected': selected,
        'courses': Course.objects.order_by('name'),
        'rooms': Room.objects.order_by('name'),
        'lesson_type_choices': Lesson.TYPE_CHOICES,
        'status_choices': Lesson.STATUS_CHOICES,
        'trial_only': False,
        'commercial_mode': True,
        'is_admin': False,
    })


@user_passes_test(_admin_check)
def lesson_agenda(request):
    target_date = _parse_date(request.GET.get('data'))
    period = _agenda_period(request.GET.get('periodo', 'hoje'))
    period_start, period_end = _agenda_period_bounds(period, target_date)
    sponte_xml_form = SponteLessonReportXMLUploadForm(initial={
        'start_date': period_start,
        'end_date': period_end,
    })
    lessons = _lesson_queryset().filter(date__gte=period_start, date__lte=period_end)
    lessons, selected = _apply_lesson_filters(request, lessons)
    lessons = _attach_lesson_display_state(list(lessons))
    selected['period'] = period
    calendar_weeks = _lesson_calendar(lessons, period_start, period_end, period)
    return render(request, 'checklists/lesson_agenda.html', {
        'title': 'Agenda de Aulas',
        'lessons': lessons,
        'calendar_weeks': calendar_weeks,
        'weekday_labels': ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom'],
        'target_date': target_date,
        'period_start': period_start,
        'period_end': period_end,
        'selected': selected,
        'courses': Course.objects.order_by('name'),
        'rooms': Room.objects.order_by('name'),
        'lesson_type_choices': Lesson.TYPE_CHOICES,
        'status_choices': Lesson.STATUS_CHOICES,
        'trial_only': False,
        'sponte_xml_form': sponte_xml_form,
        'sponte_xml_import_url': reverse('pedagogical_lessons_import_sponte_xml'),
        'is_admin': True,
    })


def _sync_lessons_from_sponte(request, *, redirect_view_name, default_period):
    target_date = _parse_date(request.POST.get('data'))
    period = _agenda_period(request.POST.get('periodo', default_period))
    redirect_to = f'{reverse(redirect_view_name)}?data={target_date.isoformat()}&periodo={period}'
    start_date, end_date = _agenda_period_bounds(period, target_date)
    job, created = enqueue_sponte_sync_job(
        kind=SponteSyncJob.KIND_SCHEDULE,
        requested_by=request.user,
        period_start=start_date,
        period_end=end_date,
        allow_past=True,
    )
    _notify_sponte_job(request, job, created)
    return redirect(redirect_to)


def _sponte_report_redirect(request, *, redirect_view_name, fallback_period='hoje'):
    target_date = _parse_date(request.POST.get('data'))
    period = _agenda_period(request.POST.get('periodo', fallback_period))
    return f'{reverse(redirect_view_name)}?data={target_date.isoformat()}&periodo={period}'


def _sponte_xml_result_message(result):
    return (
        'XML do Sponte processado. '
        f'Criadas: {result.created}; atualizadas: {result.updated}; '
        f'sem alteração: {result.unchanged}; removidas: {result.deleted}; '
        f'canceladas por ausência no XML: {result.cancelled}; ignoradas: {result.skipped}.'
    )


@require_POST
@user_passes_test(_admin_check)
def lessons_sync_sponte(request):
    return _sync_lessons_from_sponte(
        request,
        redirect_view_name='pedagogical_class_schedule',
        default_period='hoje',
    )


def _import_lessons_from_sponte_xml(request, *, redirect_view_name, fallback_period):
    redirect_to = _sponte_report_redirect(
        request,
        redirect_view_name=redirect_view_name,
        fallback_period=fallback_period,
    )
    form = SponteLessonReportXMLUploadForm(request.POST, request.FILES)
    if not form.is_valid():
        for error in form.non_field_errors():
            messages.error(request, error)
        for field_name, field_errors in form.errors.items():
            if field_name == '__all__':
                continue
            field = form.fields.get(field_name)
            label = field.label if field else field_name
            for error in field_errors:
                messages.error(request, f'{label}: {error}')
        return redirect(redirect_to)

    uploaded_xml = form.cleaned_data['report_xml']
    try:
        result = reconcile_sponte_free_class_report_xml(
            uploaded_xml.read(),
            start_date=form.cleaned_data['start_date'],
            end_date=form.cleaned_data['end_date'],
            allow_past=True,
        )
    except SponteClientError as exc:
        messages.error(request, str(exc))
        return redirect(redirect_to)

    messages.success(request, _sponte_xml_result_message(result))
    for warning in result.errors[:5]:
        messages.warning(request, warning)
    if len(result.errors) > 5:
        messages.warning(request, f'Mais {len(result.errors) - 5} aviso(s) omitidos. Verifique alunos importados do Sponte.')
    return redirect(redirect_to)


@require_POST
@user_passes_test(_admin_check)
def lessons_import_sponte_xml(request):
    return _import_lessons_from_sponte_xml(
        request,
        redirect_view_name='pedagogical_class_schedule',
        fallback_period='hoje',
    )


@require_POST
@user_passes_test(_instructor_check)
def instructor_lessons_import_sponte_xml(request):
    return _import_lessons_from_sponte_xml(
        request,
        redirect_view_name='instructor_agenda',
        fallback_period='semana',
    )


@require_POST
@user_passes_test(_commercial_operator_check)
def commercial_lessons_sync_sponte(request):
    return _sync_lessons_from_sponte(
        request,
        redirect_view_name='commercial_lesson_agenda',
        default_period='semana',
    )


@user_passes_test(_admin_check)
def trial_lessons(request):
    target_date = _parse_date(request.GET.get('data'))
    period = _agenda_period(request.GET.get('periodo', 'hoje'))
    period_start, period_end = _agenda_period_bounds(period, target_date)
    lessons = _lesson_queryset().filter(
        date__gte=period_start,
        date__lte=period_end,
        lesson_type=Lesson.TYPE_TRIAL,
    )
    lessons, selected = _apply_lesson_filters(request, lessons)
    lessons = list(lessons)
    for lesson in lessons:
        lesson.assistant_warning = lesson.assistant_warning_message()
    selected['period'] = period
    calendar_weeks = _lesson_calendar(lessons, period_start, period_end, period)
    return render(request, 'checklists/lesson_agenda.html', {
        'title': 'Aulas Experimentais ou Play',
        'lessons': lessons,
        'calendar_weeks': calendar_weeks,
        'weekday_labels': ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom'],
        'target_date': target_date,
        'period_start': period_start,
        'period_end': period_end,
        'selected': selected,
        'courses': Course.objects.order_by('name'),
        'rooms': Room.objects.order_by('name'),
        'lesson_type_choices': Lesson.TYPE_CHOICES,
        'status_choices': Lesson.STATUS_CHOICES,
        'trial_only': True,
        'is_admin': True,
    })


@user_passes_test(_admin_check)
def lesson_create(request):
    initial = {'lesson_type': Lesson.TYPE_TRIAL}
    opportunity_id = request.GET.get('oportunidade', '')
    if opportunity_id.isdigit():
        opportunity = CommercialOpportunity.objects.filter(pk=opportunity_id).first()
        if opportunity:
            initial.update({
                'commercial_opportunity': opportunity.pk,
                'student_name_snapshot': opportunity.contact_name,
                'responsible_name_snapshot': opportunity.contact_name,
                'whatsapp_snapshot': opportunity.contact_phone,
                'course': opportunity.interest_course_id,
            })
    if request.method == 'POST':
        form = LessonForm(request.POST, trial_only=True)
        if form.is_valid():
            lesson = form.save(commit=False)
            lesson.lesson_type = Lesson.TYPE_TRIAL
            lesson.created_by = request.user
            lesson.save()
            _log_change(
                request=request,
                obj=lesson,
                action='Aula pedagógica criada',
                audit_fields=LESSON_AUDIT_FIELDS,
                details=f'Aula: {lesson}; status: {lesson.get_status_display()}',
            )
            warning = lesson.assistant_warning_message()
            if warning:
                messages.warning(request, warning)
            messages.success(request, f'Aula salva com status {lesson.get_status_display()}.')
            return redirect(f'{reverse("pedagogical_class_schedule")}?data={lesson.date.isoformat()}')
    else:
        form = LessonForm(initial=initial, trial_only=True)
    return render(request, 'checklists/lesson_form.html', {
        'form': form,
        'title': 'Agendar Aula Experimental ou Play',
        'submit_label': 'Salvar aula',
        'lesson': None,
        'is_admin': True,
    })


@user_passes_test(_admin_check)
def lesson_edit(request, lesson_id):
    lesson = get_object_or_404(_lesson_queryset(), pk=lesson_id)
    if lesson.lesson_type == Lesson.TYPE_REGULAR:
        messages.error(request, 'Aulas regulares são gerenciadas pelo Sponte e ficam somente leitura no Checklist.')
        return redirect(f'{reverse("pedagogical_class_schedule")}?data={lesson.date.isoformat()}')
    previous_values_full = snapshot_instance(lesson, LESSON_AUDIT_FIELDS)
    if request.method == 'POST':
        form = LessonForm(request.POST, instance=lesson, trial_only=True)
        if form.is_valid():
            lesson = form.save()
            _log_change(
                request=request,
                obj=lesson,
                action='Aula pedagógica atualizada',
                audit_fields=LESSON_AUDIT_FIELDS,
                previous_values_full=previous_values_full,
                details=f'Aula: {lesson}; status: {lesson.get_status_display()}',
            )
            warning = lesson.assistant_warning_message()
            if warning:
                messages.warning(request, warning)
            messages.success(request, f'Aula atualizada com status {lesson.get_status_display()}.')
            return redirect(f'{reverse("pedagogical_class_schedule")}?data={lesson.date.isoformat()}')
    else:
        form = LessonForm(instance=lesson, trial_only=True)
    return render(request, 'checklists/lesson_form.html', {
        'form': form,
        'title': f'Editar aula - {lesson.student_name_snapshot}',
        'submit_label': 'Salvar alterações',
        'lesson': lesson,
        'is_admin': True,
    })
