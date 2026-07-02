import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP

from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import load_workbook

from .grafana_sync import sync_metric_area
from .models import MetricRecord, MetricType, NPSImport, NPSResponse, Position


NPS_RECORD_NOTES_PREFIX = 'Importação NPS'


@dataclass
class NPSImportResult:
    batch: NPSImport | None = None
    imported: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)
    metric_record: MetricRecord | None = None


def _normalize(value):
    text = str(value or '').strip().lower()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(char for char in text if not unicodedata.combining(char))
    return re.sub(r'\s+', ' ', text)


def _safe_json_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _as_aware(value):
    if not isinstance(value, datetime):
        return None
    if timezone.is_naive(value):
        return timezone.make_aware(value, timezone.get_current_timezone())
    return value


def _as_decimal(value):
    if value is None or value == '':
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace(',', '.')
    match = re.search(r'[-+]?\d+(?:\.\d+)?', text)
    if not match:
        return None
    return Decimal(match.group(0))


def normalize_nps_score(value):
    rating = _as_decimal(value)
    if rating is None:
        return None, None
    if rating <= 5:
        score = rating * Decimal('2')
    else:
        score = rating
    score = max(Decimal('0'), min(Decimal('10'), score))
    return rating.quantize(Decimal('0.01')), score.quantize(Decimal('0.01'))


def categorize_score(score):
    if score >= Decimal('9'):
        return NPSResponse.CATEGORY_PROMOTER
    if score >= Decimal('7'):
        return NPSResponse.CATEGORY_PASSIVE
    return NPSResponse.CATEGORY_DETRACTOR


def _find_header(headers, *needles):
    normalized_needles = [_normalize(item) for item in needles]
    for header in headers:
        normalized = _normalize(header)
        if all(needle in normalized for needle in normalized_needles):
            return header
    return None


def _ensure_nps_metric():
    position, _ = Position.objects.get_or_create(
        code='atendente-comercial',
        defaults={'name': 'Atendente Comercial', 'active': True},
    )
    metric = MetricType.objects.filter(name__icontains='NPS').order_by('id').first()
    if metric is None:
        code = slugify(f'{position.code}-6-pos-venda-pesquisa-nps')[:80] or 'nps'
        metric = MetricType.objects.create(
            code=code,
            name='Pesquisa de NPS após relatório pedagógico enviado',
            area='6 - Pós-Venda',
            position=position,
            frequency=MetricType.FREQ_MONTHLY,
            monthly_target=Decimal('85'),
            target_text='≥ 85% promotores',
            target_min=Decimal('85'),
            target_direction=MetricType.TARGET_AT_LEAST,
            unit='%',
            formula='Promotores / respostas válidas da pesquisa NPS x 100.',
            data_source='Checklist - Importação XLSX NPS',
            visualization_type=MetricType.VIZ_GAUGE,
            active=True,
        )
    else:
        changed_fields = []
        if not metric.position_id:
            metric.position = position
            changed_fields.append('position')
        if not metric.unit:
            metric.unit = '%'
            changed_fields.append('unit')
        if not metric.formula:
            metric.formula = 'Promotores / respostas válidas da pesquisa NPS x 100.'
            changed_fields.append('formula')
        if not metric.data_source:
            metric.data_source = 'Checklist - Importação XLSX NPS'
            changed_fields.append('data_source')
        if changed_fields:
            metric.save(update_fields=changed_fields)
    return metric, position


def _percent(part, total):
    if not total:
        return Decimal('0.00')
    value = Decimal(part) * Decimal('100') / Decimal(total)
    return value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


@transaction.atomic
def import_nps_xlsx(path, *, imported_by=None, source_file='', replace_existing=False, sync_grafana=True):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(value or '').strip() for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    score_header = (
        _find_header(headers, 'como voce avalia')
        or _find_header(headers, 'nps')
        or _find_header(headers, 'recomendaria')
    )
    if not score_header:
        raise ValueError('Planilha NPS sem coluna de nota reconhecida.')

    id_header = _find_header(headers, 'id')
    started_header = _find_header(headers, 'hora de inicio')
    completed_header = _find_header(headers, 'hora de conclusao')
    student_header = _find_header(headers, 'nome do(a) aluno(a)') or _find_header(headers, 'aluno')
    responsible_header = _find_header(headers, 'nome do(a) responsavel') or _find_header(headers, 'responsavel')
    comment_header = _find_header(headers, 'o que voce mais gostou') or headers[-1]

    if replace_existing:
        NPSImport.objects.all().delete()
        metric_ids = list(MetricType.objects.filter(name__icontains='NPS').values_list('id', flat=True))
        if metric_ids:
            MetricRecord.objects.filter(metric_id__in=metric_ids, notes__startswith=NPS_RECORD_NOTES_PREFIX).delete()

    result = NPSImportResult()
    batch = NPSImport.objects.create(source_file=source_file or str(path), imported_by=imported_by)
    result.batch = batch

    responses = []
    completed_dates = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = {header: value for header, value in zip(headers, row)}
        original_rating, score = normalize_nps_score(data.get(score_header))
        if score is None:
            result.skipped += 1
            result.errors.append(f'Linha {row_number}: nota NPS vazia ou inválida.')
            continue

        completed_at = _as_aware(data.get(completed_header))
        started_at = _as_aware(data.get(started_header))
        if completed_at:
            completed_dates.append(completed_at.date())

        category = categorize_score(score)
        raw_data = {key: _safe_json_value(value) for key, value in data.items()}
        responses.append(NPSResponse(
            import_batch=batch,
            external_id=str(data.get(id_header) or '').strip() if id_header else '',
            started_at=started_at,
            completed_at=completed_at,
            student_name=str(data.get(student_header) or '').strip() if student_header else '',
            responsible_name=str(data.get(responsible_header) or '').strip() if responsible_header else '',
            original_rating=original_rating,
            score_0_10=score,
            category=category,
            comment=str(data.get(comment_header) or '').strip() if comment_header else '',
            raw_data=raw_data,
        ))

    if responses:
        NPSResponse.objects.bulk_create(responses)

    promoters = sum(1 for item in responses if item.category == NPSResponse.CATEGORY_PROMOTER)
    passives = sum(1 for item in responses if item.category == NPSResponse.CATEGORY_PASSIVE)
    detractors = sum(1 for item in responses if item.category == NPSResponse.CATEGORY_DETRACTOR)
    total = len(responses)
    promoter_percent = _percent(promoters, total)
    detractor_percent = _percent(detractors, total)
    nps_score = (promoter_percent - detractor_percent).quantize(Decimal('0.01'))

    batch.total_responses = total
    batch.promoters = promoters
    batch.passives = passives
    batch.detractors = detractors
    batch.promoter_percent = promoter_percent
    batch.nps_score = nps_score
    if completed_dates:
        batch.period_start = min(completed_dates)
        batch.period_end = max(completed_dates)
    batch.save()

    metric, position = _ensure_nps_metric()
    record_date = batch.period_end or timezone.localdate()
    MetricRecord.objects.filter(
        metric=metric,
        position=position,
        date=record_date,
        notes__startswith=NPS_RECORD_NOTES_PREFIX,
    ).delete()
    metric_record = MetricRecord.objects.create(
        metric=metric,
        position=position,
        date=record_date,
        value=promoter_percent,
        notes=(
            f'{NPS_RECORD_NOTES_PREFIX}: {batch.source_file}; '
            f'{promoters} promotores, {passives} passivos, {detractors} detratores; '
            f'NPS líquido {nps_score}.'
        ),
        created_by=imported_by,
    )
    batch.metric_record = metric_record
    batch.save(update_fields=['metric_record'])
    result.metric_record = metric_record
    result.imported = total

    if sync_grafana:
        sync_metric_area(metric.area)

    return result
