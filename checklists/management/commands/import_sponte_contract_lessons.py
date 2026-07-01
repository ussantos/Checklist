from collections import defaultdict, deque
from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path
import time
import xml.etree.ElementTree as ET

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from checklists.management.commands.import_lesson_feedbacks_xlsx import (
    Command as FeedbackWorkbookCommand,
    ImportSummary as FeedbackWorkbookSummary,
)
from checklists.models import Lesson, LessonFeedback, PedagogicalStudent
from checklists.services import (
    create_pedagogical_report_task_for_feedback,
    create_qualified_sale_opportunity_for_lesson_feedback,
)
from checklists.sponte import (
    SponteClientError,
    SponteConfigurationError,
    _child_text,
    _children_text_map,
    _course_for_schedule,
    _is_not_found_message,
    _is_success_message,
    _lesson_status_from_sponte,
    _lesson_status_label_from_sponte_fields,
    _local_name,
    _parse_sponte_date,
    _parse_sponte_time,
    _room_for_schedule,
    _truncate,
    import_sponte_students,
    parse_sponte_free_class_schedule,
)
from checklists.sponte_client import SponteAPIError, SponteSOAPClient


@dataclass
class ContractDiscipline:
    external_id: str = ''
    name: str = ''
    module: int | None = None


@dataclass
class StudentContract:
    contract_id: str = ''
    free_contract_id: str = ''
    number: str = ''
    status: str = ''
    course_external_id: str = ''
    course_name: str = ''
    disciplines: list[ContractDiscipline] = field(default_factory=list)


@dataclass
class ImportedLesson:
    lesson: Lesson
    status_from_sponte: bool
    status_label: str


@dataclass
class HistoricalImportSummary:
    students_imported_created: int = 0
    students_imported_updated: int = 0
    students_imported_unchanged: int = 0
    students_processed: int = 0
    contracts_read: int = 0
    months_processed: int = 0
    api_calls: int = 0
    agenda_records: int = 0
    diary_records: int = 0
    lessons_created: int = 0
    lessons_updated: int = 0
    lessons_unchanged: int = 0
    feedback_rows_ready: int = 0
    feedback_rows_already_present: int = 0
    done_lessons_with_feedback: int = 0
    feedbacks_created: int = 0
    feedbacks_updated: int = 0
    feedbacks_unchanged: int = 0
    report_tasks_created: int = 0
    qualified_opportunities_created: int = 0
    pending_feedback_lessons: int = 0
    unmatched_feedback_rows: int = 0
    skipped_records: int = 0
    api_errors: list[str] = field(default_factory=list)
    row_errors: list[str] = field(default_factory=list)
    output_path: str = ''


def _month_chunks(start_date, end_date):
    cursor = date(start_date.year, start_date.month, 1)
    while cursor <= end_date:
        next_month = (cursor.replace(day=28) + timedelta(days=4)).replace(day=1)
        month_start = max(cursor, start_date)
        month_end = min(next_month - timedelta(days=1), end_date)
        yield month_start, month_end
        cursor = next_month


def _parse_int(value):
    value = (value or '').strip()
    if not value:
        return None
    try:
        return int(value)
    except ValueError:
        return None


def _status_label(status):
    return dict(Lesson.STATUS_CHOICES).get(status, status)


class Command(BaseCommand):
    help = (
        'Importa aulas historicas do Sponte lendo contratos de Aulas Livres, '
        'consulta o diario para status e cruza com a planilha de feedbacks.'
    )

    def add_arguments(self, parser):
        parser.add_argument('feedback_xlsx', help='Caminho da planilha XLSX de feedbacks.')
        parser.add_argument('--start-date', default='2025-11-01', help='Data inicial YYYY-MM-DD.')
        parser.add_argument('--end-date', default='2026-06-30', help='Data final YYYY-MM-DD.')
        parser.add_argument('--dry-run', action='store_true', help='Consulta e analisa sem gravar no banco.')
        parser.add_argument(
            '--student-external-id',
            action='append',
            default=[],
            help='Limita a um ou mais AlunoID do Sponte para teste.',
        )
        parser.add_argument(
            '--skip-student-import',
            action='store_true',
            help='Não executa a importação de alunos do Sponte antes da agenda.',
        )
        parser.add_argument(
            '--output',
            default='',
            help='Caminho do XLSX de aulas com feedback pendente. Padrão: backups/exports/aulas_sponte_sem_feedback_<timestamp>.xlsx',
        )
        parser.add_argument(
            '--fallback-user',
            default='',
            help='Usuário usado como autor quando o professor da planilha não existir no sistema.',
        )
        parser.add_argument(
            '--pause-every',
            type=int,
            default=25,
            help='Quantidade de chamadas ao Sponte antes de pausar.',
        )
        parser.add_argument(
            '--pause-seconds',
            type=int,
            default=65,
            help='Tempo de pausa entre lotes de chamadas ao Sponte.',
        )

    def handle(self, *args, **options):
        feedback_path = Path(options['feedback_xlsx'])
        if not feedback_path.exists():
            raise CommandError(f'Arquivo não encontrado: {feedback_path}')
        try:
            start_date = date.fromisoformat(options['start_date'])
            end_date = date.fromisoformat(options['end_date'])
        except ValueError as exc:
            raise CommandError('Use datas no formato YYYY-MM-DD.') from exc
        if end_date < start_date:
            raise CommandError('A data final deve ser maior ou igual à data inicial.')

        summary = HistoricalImportSummary()
        self._api_calls_since_pause = 0

        if not options['skip_student_import']:
            self._import_students(summary)

        feedback_parser = FeedbackWorkbookCommand()
        feedback_rows, feedback_summary = self._load_feedback_rows(feedback_parser, feedback_path, start_date, end_date)
        summary.feedback_rows_ready = len(feedback_rows)
        rows_by_student_date = self._feedback_rows_by_student_date(feedback_rows, summary)
        users = feedback_parser._user_lookup()
        fallback_user = feedback_parser._fallback_user(options['fallback_user'])

        students = self._students(options['student_external_id'])
        client = SponteSOAPClient()
        pending_lessons = []
        imported_lessons = []

        for student in students:
            summary.students_processed += 1
            self.stdout.write(f'{student.name} ({student.external_id}): lendo contratos...')
            contracts = self._fetch_contracts(client, student, summary, options)
            summary.contracts_read += len(contracts)
            if not contracts:
                self.stdout.write(self.style.WARNING(f'{student.name}: nenhum contrato de Aulas Livres encontrado.'))

            for month_start, month_end in _month_chunks(start_date, end_date):
                summary.months_processed += 1
                agenda_records = self._fetch_agenda(client, student, month_start, month_end, summary, options)
                if not agenda_records:
                    continue
                self.stdout.write(f'  {month_start:%m/%Y}: {len(agenda_records)} aula(s) na agenda.')
                for agenda_record in agenda_records:
                    record = self._record_with_diary_status(client, student, agenda_record, contracts, summary, options)
                    if options['dry_run']:
                        continue
                    try:
                        imported = self._upsert_historical_lesson(student, record, summary)
                    except Exception as exc:  # noqa: BLE001 - command must continue with other Sponte rows.
                        summary.row_errors.append(f'{student.name} {record.get("date")}: {exc}')
                        continue
                    if imported is None:
                        continue
                    imported_lessons.append(imported)
                    if imported.lesson.status == Lesson.STATUS_DONE:
                        self._ensure_feedback(
                            imported.lesson,
                            rows_by_student_date,
                            feedback_parser,
                            users,
                            fallback_user,
                            pending_lessons,
                            summary,
                        )

        unmatched_rows = self._remaining_feedback_rows(rows_by_student_date)
        summary.unmatched_feedback_rows = len(unmatched_rows)
        summary.pending_feedback_lessons = len(pending_lessons)
        if not options['dry_run']:
            output_path = self._output_path(options['output'])
            self._write_pending_workbook(output_path, pending_lessons, unmatched_rows)
            summary.output_path = str(output_path)

        self._print_summary(summary, feedback_summary, dry_run=options['dry_run'])
        if imported_lessons:
            self.stdout.write(f'Aulas importadas/analisadas: {len(imported_lessons)}')

    def _import_students(self, summary):
        try:
            result = import_sponte_students()
        except (SponteConfigurationError, SponteClientError) as exc:
            summary.api_errors.append(f'Importação inicial de alunos: {exc}')
            return
        summary.students_imported_created = result.created
        summary.students_imported_updated = result.updated
        summary.students_imported_unchanged = result.unchanged
        if result.errors:
            summary.api_errors.extend(f'Alunos Sponte: {error}' for error in result.errors)

    def _load_feedback_rows(self, parser, feedback_path, start_date, end_date):
        workbook = load_workbook(feedback_path, read_only=True, data_only=True)
        feedback_summary = FeedbackWorkbookSummary()
        student_lookup = parser._student_lookup()
        rows = parser._parse_workbook(workbook, student_lookup, start_date, end_date, feedback_summary)
        rows.sort(key=lambda row: (row.student.name, row.class_date, row.row_number))
        return rows, feedback_summary

    def _feedback_rows_by_student_date(self, rows, summary):
        existing_keys = set(
            LessonFeedback.objects
            .filter(lesson__date__gte=min((row.class_date for row in rows), default=date.max))
            .filter(lesson__date__lte=max((row.class_date for row in rows), default=date.min))
            .values_list('lesson__student_id', 'lesson__date', 'module_number', 'lesson_number')
        )
        rows_by_student_date = defaultdict(deque)
        for row in rows:
            key = (row.student.id, row.class_date, row.module_number, row.lesson_number)
            if key in existing_keys:
                summary.feedback_rows_already_present += 1
                continue
            rows_by_student_date[(row.student.id, row.class_date)].append(row)
        return rows_by_student_date

    def _students(self, external_ids):
        queryset = (
            PedagogicalStudent.objects
            .exclude(external_id='')
            .order_by('name', 'id')
        )
        if external_ids:
            queryset = queryset.filter(external_id__in={str(value).strip() for value in external_ids if str(value).strip()})
        students = list(queryset)
        if not students:
            raise CommandError('Nenhum aluno com ID externo do Sponte foi encontrado no Checklist.')
        return students

    def _call_sponte(self, client, operation, data, summary, options):
        try:
            response = client.call(operation, data, use_cache=False)
        finally:
            summary.api_calls += 1
            self._api_calls_since_pause += 1
            if options['pause_every'] > 0 and self._api_calls_since_pause >= options['pause_every']:
                time.sleep(max(options['pause_seconds'], 0))
                self._api_calls_since_pause = 0
        return response

    def _fetch_contracts(self, client, student, summary, options):
        try:
            response = self._call_sponte(
                client,
                'GetMatriculas',
                {'sParametrosBusca': f'AlunoID={student.external_id}'},
                summary,
                options,
            )
        except SponteAPIError as exc:
            summary.api_errors.append(f'{student.name}: contratos: {exc}')
            return []
        try:
            return self._parse_contracts(response.xml_text)
        except SponteClientError as exc:
            summary.api_errors.append(f'{student.name}: contratos: {exc}')
            return []

    def _parse_contracts(self, xml_text):
        try:
            root = ET.fromstring(xml_text)
        except ET.ParseError as exc:
            raise SponteClientError(f'Resposta XML inválida do Sponte: {exc}') from exc
        contracts = []
        messages = []
        for item in root.iter():
            if _local_name(item.tag) != 'wsMatricula':
                continue
            operation_return = _child_text(item, 'RetornoOperacao')
            if operation_return:
                messages.append(operation_return)
            free_contract_id = _child_text(item, 'ContratoAulaLivreID')
            if not free_contract_id:
                continue
            contract = StudentContract(
                contract_id=_child_text(item, 'ContratoID'),
                free_contract_id=free_contract_id,
                number=_child_text(item, 'NumeroContrato'),
                status=_child_text(item, 'Situacao'),
                course_external_id=_child_text(item, 'CursoID'),
                course_name=_child_text(item, 'NomeCurso'),
            )
            for discipline in item.iter():
                if _local_name(discipline.tag) != 'wsDisciplinas':
                    continue
                contract.disciplines.append(ContractDiscipline(
                    external_id=_child_text(discipline, 'DisciplinaID'),
                    name=_child_text(discipline, 'Nome'),
                    module=_parse_int(_child_text(discipline, 'Modulo')),
                ))
            contracts.append(contract)
        if not contracts and messages:
            unexpected = [
                message for message in sorted(set(messages))
                if not (_is_success_message(message) or _is_not_found_message(message))
            ]
            if unexpected:
                raise SponteClientError('; '.join(unexpected))
        return contracts

    def _fetch_agenda(self, client, student, start_date, end_date, summary, options):
        try:
            response = self._call_sponte(
                client,
                'GetAgendaAluno',
                {
                    'nAlunoID': student.external_id,
                    'sTurmaID': '',
                    'sCursoID': '',
                    'dDataInicio': start_date.strftime('%d/%m/%Y'),
                    'dDataTermino': end_date.strftime('%d/%m/%Y'),
                },
                summary,
                options,
            )
            records = parse_sponte_free_class_schedule(response.xml_text, student_external_id=student.external_id)
        except (SponteAPIError, SponteClientError) as exc:
            summary.api_errors.append(f'{student.name}: agenda {start_date:%m/%Y}: {exc}')
            return []
        records = [record for record in records if self._record_in_range(record, start_date, end_date)]
        summary.agenda_records += len(records)
        return records

    def _record_in_range(self, record, start_date, end_date):
        class_date = _parse_sponte_date(record.get('date'))
        return bool(class_date and start_date <= class_date <= end_date)

    def _record_with_diary_status(self, client, student, agenda_record, contracts, summary, options):
        diary_records = self._fetch_diary_for_agenda_record(client, student, agenda_record, contracts, summary, options)
        diary_record = self._matching_diary_record(agenda_record, diary_records)
        if not diary_record:
            return agenda_record
        merged = {**agenda_record}
        for field_name, value in diary_record.items():
            if value or field_name in {'lesson_status_from_sponte'}:
                merged[field_name] = value
        return merged

    def _fetch_diary_for_agenda_record(self, client, student, agenda_record, contracts, summary, options):
        class_date = _parse_sponte_date(agenda_record.get('date'))
        course_id = (agenda_record.get('course_external_id') or '').strip()
        discipline_id = (agenda_record.get('discipline_external_id') or '').strip()
        if not (class_date and course_id and discipline_id):
            return []
        records = []
        for module in self._candidate_modules(contracts, course_id, discipline_id):
            try:
                response = self._call_sponte(
                    client,
                    'GetDiarioAulasLivres',
                    {
                        'nAlunoID': student.external_id,
                        'nCursoID': course_id,
                        'nDisciplinaID': discipline_id,
                        'dDataInicio': class_date.strftime('%d/%m/%Y'),
                        'dDataTermino': class_date.strftime('%d/%m/%Y'),
                        'nModulo': module,
                        'sParametrosBusca': '',
                    },
                    summary,
                    options,
                )
            except SponteAPIError as exc:
                summary.api_errors.append(f'{student.name}: diário {class_date:%d/%m/%Y}: {exc}')
                continue
            parsed = self._parse_diary_records(response.xml_text, student.external_id)
            if parsed:
                records.extend(parsed)
                break
        summary.diary_records += len(records)
        return records

    def _candidate_modules(self, contracts, course_id, discipline_id):
        candidates = []
        for contract in contracts:
            if contract.course_external_id and contract.course_external_id != course_id:
                continue
            for discipline in contract.disciplines:
                if discipline.external_id == discipline_id and discipline.module is not None:
                    candidates.append(discipline.module)
        candidates.extend([0, 1, 2, 3])
        seen = set()
        unique = []
        for candidate in candidates:
            if candidate in seen:
                continue
            seen.add(candidate)
            unique.append(candidate)
        return unique

    def _parse_diary_records(self, xml_text, student_external_id):
        try:
            root = ET.fromstring(xml_text)
        except ET.ParseError as exc:
            raise SponteClientError(f'Resposta XML inválida do Sponte: {exc}') from exc

        records = []
        for diary in root.iter():
            if _local_name(diary.tag) != 'wsDiarioAulasLivres':
                continue
            student_id = _child_text(diary, 'AlunoID') or student_external_id
            course_name = _child_text(diary, 'Curso')
            course_external_id = _child_text(diary, 'CursoID')
            discipline_name = _child_text(diary, 'Disciplina')
            discipline_external_id = _child_text(diary, 'DisciplinaID')
            contract_number = _child_text(diary, 'NumeroContrato')
            for item in diary.iter():
                if _local_name(item.tag) != 'wsAulasLivreDiario':
                    continue
                fields = _children_text_map(item)
                status_label = _lesson_status_label_from_sponte_fields(fields)
                records.append({
                    'student_external_id': student_id,
                    'free_class_id': _child_text(item, 'AulaLivreID'),
                    'date': _child_text(item, 'DataAula'),
                    'start_time': _child_text(item, 'HorarioInicial'),
                    'end_time': _child_text(item, 'HorarioFinal'),
                    'lesson_status': status_label or _status_label(Lesson.STATUS_NOT_GIVEN),
                    'lesson_status_raw': _child_text(item, 'Situacao'),
                    'lesson_status_from_sponte': bool(status_label),
                    'course_name': course_name or discipline_name or 'Curso Sponte',
                    'room_name': _child_text(item, 'Sala'),
                    'teacher_name': _child_text(item, 'Professor'),
                    'course_external_id': course_external_id,
                    'discipline_external_id': discipline_external_id,
                    'discipline_name': discipline_name,
                    'teacher_external_id': _child_text(item, 'ProfessorID'),
                    'contract_number': contract_number,
                })
        return records

    def _matching_diary_record(self, agenda_record, diary_records):
        if not diary_records:
            return None
        free_class_id = (agenda_record.get('free_class_id') or '').strip()
        if free_class_id:
            for record in diary_records:
                if (record.get('free_class_id') or '').strip() == free_class_id:
                    return record
        agenda_date = _parse_sponte_date(agenda_record.get('date'))
        agenda_start = _parse_sponte_time(agenda_record.get('start_time'))
        agenda_end = _parse_sponte_time(agenda_record.get('end_time'))
        for record in diary_records:
            if (
                _parse_sponte_date(record.get('date')) == agenda_date
                and _parse_sponte_time(record.get('start_time')) == agenda_start
                and _parse_sponte_time(record.get('end_time')) == agenda_end
            ):
                return record
        return diary_records[0] if len(diary_records) == 1 else None

    def _upsert_historical_lesson(self, student, record, summary):
        class_date = _parse_sponte_date(record.get('date'))
        start_time = _parse_sponte_time(record.get('start_time'))
        end_time = _parse_sponte_time(record.get('end_time'))
        free_class_id = _truncate(record.get('free_class_id'), 80)
        if not all([class_date, start_time, end_time, free_class_id]):
            summary.skipped_records += 1
            return None

        external_id = f'aula_livre:{student.external_id}:{free_class_id}:{class_date.isoformat()}'
        course = _course_for_schedule(record)
        room = _room_for_schedule(record)
        status_from_sponte = bool(record.get('lesson_status_from_sponte'))
        incoming_status = _lesson_status_from_sponte(record.get('lesson_status'))
        status_label = record.get('lesson_status') or _status_label(incoming_status)
        now = timezone.now()
        status_note = (
            f'Situação no Sponte: {status_label}'
            if status_from_sponte
            else 'Situação no Sponte: não informada pelo diário; padrão Checklist Não dada.'
        )
        notes = '\n'.join([
            'Aula Livre histórica sincronizada do Sponte.',
            f'Professor: {record.get("teacher_name") or "-"}',
            status_note,
            f'CursoID: {record.get("course_external_id") or "-"}',
            f'DisciplinaID: {record.get("discipline_external_id") or "-"}',
            f'Contrato: {record.get("contract_number") or "-"}',
        ])
        values = {
            'student': student,
            'student_name_snapshot': student.name,
            'responsible_name_snapshot': student.responsible_name,
            'whatsapp_snapshot': student.whatsapp,
            'lesson_type': Lesson.TYPE_REGULAR,
            'course': course,
            'room': room,
            'date': class_date,
            'start_time': start_time,
            'end_time': end_time,
            'notes': notes,
            'source': Lesson.SOURCE_SPONTE,
            'external_id': external_id,
            'synced_at': now,
        }
        lesson = Lesson.objects.filter(source=Lesson.SOURCE_SPONTE, external_id=external_id).first()
        if lesson is None:
            values['status'] = incoming_status if status_from_sponte else Lesson.STATUS_NOT_GIVEN
            lesson = Lesson.objects.create(**values)
            summary.lessons_created += 1
            return ImportedLesson(lesson=lesson, status_from_sponte=status_from_sponte, status_label=status_label)

        if status_from_sponte:
            values['status'] = incoming_status
        changed = False
        for field_name, value in values.items():
            if getattr(lesson, field_name) != value:
                setattr(lesson, field_name, value)
                changed = True
        if changed:
            lesson.save(update_fields=[*values.keys(), 'updated_at'])
            summary.lessons_updated += 1
        else:
            summary.lessons_unchanged += 1
        return ImportedLesson(lesson=lesson, status_from_sponte=status_from_sponte, status_label=status_label)

    def _ensure_feedback(
        self,
        lesson,
        rows_by_student_date,
        feedback_parser,
        users,
        fallback_user,
        pending_lessons,
        summary,
    ):
        if LessonFeedback.objects.filter(lesson=lesson).exists():
            summary.done_lessons_with_feedback += 1
            return
        row = self._next_feedback_row(lesson, rows_by_student_date)
        if row is None:
            pending_lessons.append(lesson)
            return
        actor = feedback_parser._actor_for_teacher(row.teacher_name, users, fallback_user)
        try:
            with transaction.atomic():
                feedback, created, updated = feedback_parser._upsert_feedback(row, lesson, actor)
                report_task, report_created = create_pedagogical_report_task_for_feedback(feedback, actor=actor)
                opportunity, opportunity_created = create_qualified_sale_opportunity_for_lesson_feedback(feedback, actor=actor)
        except Exception as exc:  # noqa: BLE001 - report and continue.
            summary.row_errors.append(f'{row.sheet_name} linha {row.row_number}: {exc}')
            pending_lessons.append(lesson)
            return
        summary.feedbacks_created += int(created)
        summary.feedbacks_updated += int(updated)
        summary.feedbacks_unchanged += int(not created and not updated)
        summary.report_tasks_created += int(bool(report_task and report_created))
        summary.qualified_opportunities_created += int(bool(opportunity and opportunity_created))

    def _next_feedback_row(self, lesson, rows_by_student_date):
        rows = rows_by_student_date.get((lesson.student_id, lesson.date))
        if not rows:
            return None
        return rows.popleft()

    def _remaining_feedback_rows(self, rows_by_student_date):
        rows = []
        for queue in rows_by_student_date.values():
            rows.extend(list(queue))
        return rows

    def _output_path(self, output):
        if output:
            path = Path(output)
        else:
            timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
            path = Path('backups') / 'exports' / f'aulas_sponte_sem_feedback_{timestamp}.xlsx'
        path.parent.mkdir(parents=True, exist_ok=True)
        return path

    def _write_pending_workbook(self, output_path, pending_lessons, unmatched_rows):
        workbook = Workbook()
        pending_sheet = workbook.active
        pending_sheet.title = 'Aulas sem feedback'
        self._write_pending_lessons_sheet(pending_sheet, pending_lessons)
        unmatched_sheet = workbook.create_sheet('Feedbacks sem aula')
        self._write_unmatched_feedback_sheet(unmatched_sheet, unmatched_rows)
        workbook.save(output_path)

    def _write_pending_lessons_sheet(self, sheet, lessons):
        headers = [
            'Aluno', 'Status do aluno', 'Data', 'Início', 'Fim', 'Curso', 'Sala',
            'Situação Sponte', 'ID aula Sponte', 'Observação',
        ]
        rows = []
        for lesson in lessons:
            rows.append([
                lesson.student.name if lesson.student_id else lesson.student_name_snapshot,
                lesson.student.get_status_display() if lesson.student_id else '',
                lesson.date,
                lesson.start_time,
                lesson.end_time,
                lesson.course.name if lesson.course_id else '',
                lesson.room.name if lesson.room_id else '',
                lesson.get_status_display(),
                lesson.external_id,
                'Aula com Presença no Sponte sem feedback correspondente na planilha.',
            ])
        self._write_sheet(sheet, headers, rows)

    def _write_unmatched_feedback_sheet(self, sheet, rows):
        headers = [
            'Aba', 'Linha', 'Aluno', 'Data', 'Curso na planilha', 'Módulo', 'Aula',
            'Professor', 'Observação',
        ]
        data = []
        for row in rows:
            data.append([
                row.sheet_name,
                row.row_number,
                row.student.name,
                row.class_date,
                row.course_hint,
                row.module_number,
                row.lesson_number,
                row.teacher_name,
                'Feedback da planilha não foi usado por falta de aula com Presença correspondente.',
            ])
        self._write_sheet(sheet, headers, data)

    def _write_sheet(self, sheet, headers, rows):
        header_fill = PatternFill('solid', fgColor='1F4E78')
        header_font = Font(bold=True, color='FFFFFF')
        thin = Side(style='thin', color='D9E2F3')
        border = Border(bottom=thin)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        for row in rows:
            sheet.append(row)
        sheet.freeze_panes = 'A2'
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max(len(str(cell.value or '')) for cell in column_cells)
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 55)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = border
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if hasattr(cell.value, 'strftime'):
                    if cell.value.__class__.__name__ == 'date':
                        cell.number_format = 'dd/mm/yyyy'
                    else:
                        cell.number_format = 'hh:mm'

    def _print_summary(self, summary, feedback_summary, *, dry_run):
        label = 'Diagnóstico' if dry_run else 'Importação histórica'
        self.stdout.write(self.style.SUCCESS(f'{label} finalizado.'))
        for name, value in [
            ('students_imported_created', summary.students_imported_created),
            ('students_imported_updated', summary.students_imported_updated),
            ('students_imported_unchanged', summary.students_imported_unchanged),
            ('students_processed', summary.students_processed),
            ('contracts_read', summary.contracts_read),
            ('months_processed', summary.months_processed),
            ('api_calls', summary.api_calls),
            ('agenda_records', summary.agenda_records),
            ('diary_records', summary.diary_records),
            ('lessons_created', summary.lessons_created),
            ('lessons_updated', summary.lessons_updated),
            ('lessons_unchanged', summary.lessons_unchanged),
            ('feedback_rows_in_period', feedback_summary.rows_in_period),
            ('feedback_rows_ready', summary.feedback_rows_ready),
            ('feedback_rows_already_present', summary.feedback_rows_already_present),
            ('done_lessons_with_feedback', summary.done_lessons_with_feedback),
            ('feedbacks_created', summary.feedbacks_created),
            ('feedbacks_updated', summary.feedbacks_updated),
            ('feedbacks_unchanged', summary.feedbacks_unchanged),
            ('report_tasks_created', summary.report_tasks_created),
            ('qualified_opportunities_created', summary.qualified_opportunities_created),
            ('pending_feedback_lessons', summary.pending_feedback_lessons),
            ('unmatched_feedback_rows', summary.unmatched_feedback_rows),
            ('skipped_records', summary.skipped_records),
        ]:
            self.stdout.write(f'{name}: {value}')
        if summary.output_path:
            self.stdout.write(f'xlsx_pendencias: {summary.output_path}')
        for error in summary.api_errors[:20]:
            self.stdout.write(self.style.WARNING(f'API: {error}'))
        for error in summary.row_errors[:20]:
            self.stdout.write(self.style.WARNING(f'Linha: {error}'))
        if len(summary.api_errors) > 20 or len(summary.row_errors) > 20:
            self.stdout.write(self.style.WARNING('Há mais erros omitidos no resumo.'))
