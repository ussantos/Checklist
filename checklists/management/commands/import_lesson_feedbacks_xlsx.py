import re
import time
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from checklists.models import Lesson, LessonFeedback, PedagogicalStudent
from checklists.services import (
    create_pedagogical_report_task_for_feedback,
    create_post_sale_opportunity_for_lesson_feedback,
)
from checklists.sponte import (
    _course_for_schedule,
    _parse_sponte_date,
    _parse_sponte_time,
    _room_for_schedule,
    parse_sponte_free_class_schedule,
)
from checklists.sponte_client import SponteAPIError, SponteSOAPClient


PROGRAMMING_EMPTY_VALUES = {'', '-', 'n/a', 'na', 'não se aplica', 'nao se aplica'}


@dataclass
class FeedbackImportRow:
    sheet_name: str
    row_number: int
    student: PedagogicalStudent
    class_date: date
    module_number: int
    lesson_number: int
    course_hint: str
    teacher_name: str
    punctuality_score: int
    assembly_comment: str
    assembly_score: int
    has_programming: bool
    programming_comment: str
    programming_score: int | None
    participation_comment: str
    participation_score: int
    behavior_comment: str
    behavior_score: int
    general_comment: str


@dataclass
class ScheduleMatch:
    record: dict | None
    lesson: Lesson | None = None


@dataclass
class ImportSummary:
    rows_in_period: int = 0
    rows_ready: int = 0
    rows_imported: int = 0
    skipped_unmatched_student: int = 0
    skipped_incomplete: int = 0
    skipped_no_schedule: int = 0
    skipped_ambiguous_schedule: int = 0
    lessons_created: int = 0
    lessons_updated: int = 0
    feedbacks_created: int = 0
    feedbacks_updated: int = 0
    feedbacks_unchanged: int = 0
    report_tasks_created: int = 0
    post_sale_opportunities_created: int = 0
    api_errors: list[str] = field(default_factory=list)
    row_errors: list[str] = field(default_factory=list)


def _normalize(value):
    value = '' if value is None else str(value)
    value = unicodedata.normalize('NFKD', value)
    value = ''.join(char for char in value if not unicodedata.combining(char))
    value = value.lower()
    value = re.sub(r'[^a-z0-9]+', ' ', value)
    return re.sub(r'\s+', ' ', value).strip()


def _sheet_student_name(sheet_name):
    if '-' in sheet_name:
        return sheet_name.split('-', 1)[1].strip()
    return sheet_name.strip()


def _parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = (str(value).strip() if value is not None else '')
    if not text:
        return None
    match = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', text)
    if not match:
        return None
    day, month, year = (int(part) for part in match.groups())
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _parse_int(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)) and int(value) == value:
        return int(value)
    text = str(value).strip()
    match = re.search(r'\d+', text)
    if not match:
        return None
    return int(match.group())


def _parse_score(value):
    if value is None:
        return None
    text = str(value).strip()
    if _normalize(text) in PROGRAMMING_EMPTY_VALUES:
        return None
    try:
        number = Decimal(text.replace(',', '.')) if isinstance(value, str) else Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
    score = int(number.quantize(Decimal('1'), rounding=ROUND_HALF_UP))
    if score < 0 or score > 10:
        return None
    return score


def _parse_punctuality(value):
    text = _normalize(value)
    if text in {'sim', 's', 'presenca', 'pontual'}:
        return 10
    if text in {'nao', 'n', 'não', 'atraso'}:
        return 5
    score = _parse_score(value)
    if score in {5, 10}:
        return score
    return None


def _text(value):
    return '' if value is None else str(value).strip()


def _is_programming_empty(comment, score):
    return _normalize(comment) in PROGRAMMING_EMPTY_VALUES and score is None


def _header_index(header_keys, *candidates):
    candidate_keys = {_normalize(candidate) for candidate in candidates}
    for index, key in enumerate(header_keys):
        if key in candidate_keys:
            return index
    return None


def _header_indices(header_keys, *candidates):
    candidate_keys = {_normalize(candidate) for candidate in candidates}
    return [index for index, key in enumerate(header_keys) if key in candidate_keys]


def _header_index_contains(header_keys, *, include, exclude=()):
    include = tuple(_normalize(value) for value in include)
    exclude = tuple(_normalize(value) for value in exclude)
    for index, key in enumerate(header_keys):
        if all(part in key for part in include) and not any(part in key for part in exclude):
            return index
    return None


def _value(row, index):
    if index is None or index >= len(row):
        return None
    return row[index]


def _date_chunks(start_date, end_date, days=89):
    cursor = start_date
    while cursor <= end_date:
        chunk_end = min(cursor + timedelta(days=days), end_date)
        yield cursor, chunk_end
        cursor = chunk_end + timedelta(days=1)


class Command(BaseCommand):
    help = 'Importa feedbacks de aula de uma planilha XLSX e vincula às aulas históricas do Sponte.'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', help='Caminho da planilha XLSX com os feedbacks.')
        parser.add_argument('--start-date', default='2025-11-01', help='Data inicial no formato YYYY-MM-DD.')
        parser.add_argument('--end-date', default='2026-06-28', help='Data final no formato YYYY-MM-DD.')
        parser.add_argument('--dry-run', action='store_true', help='Analisa sem gravar alterações.')
        parser.add_argument(
            '--existing-lessons-only',
            action='store_true',
            help='Não consulta o Sponte; usa somente aulas já existentes no banco.',
        )
        parser.add_argument(
            '--fallback-user',
            default='',
            help='Usuário usado como autor quando o professor da planilha não existir no sistema.',
        )
        parser.add_argument(
            '--pause-every',
            type=int,
            default=25,
            help='Quantidade de chamadas ao Sponte antes de pausar para respeitar limite de API.',
        )
        parser.add_argument(
            '--pause-seconds',
            type=int,
            default=65,
            help='Tempo de pausa entre lotes de chamadas ao Sponte.',
        )

    def handle(self, *args, **options):
        xlsx_path = Path(options['xlsx_path'])
        if not xlsx_path.exists():
            raise CommandError(f'Arquivo não encontrado: {xlsx_path}')

        try:
            start_date = date.fromisoformat(options['start_date'])
            end_date = date.fromisoformat(options['end_date'])
        except ValueError as exc:
            raise CommandError('Use datas no formato YYYY-MM-DD.') from exc
        if end_date < start_date:
            raise CommandError('A data final deve ser maior ou igual à data inicial.')

        summary = ImportSummary()
        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
        students = self._student_lookup()
        users = self._user_lookup()
        fallback_user = self._fallback_user(options['fallback_user'])

        rows = self._parse_workbook(workbook, students, start_date, end_date, summary)
        summary.rows_ready = len(rows)
        self.stdout.write(f'Linhas no período: {summary.rows_in_period}')
        self.stdout.write(f'Linhas prontas para importação: {summary.rows_ready}')

        schedule_index = self._existing_lesson_index(rows)
        if not options['existing_lessons_only']:
            self._load_sponte_schedule(rows, schedule_index, start_date, end_date, summary, options)

        for row in rows:
            actor = self._actor_for_teacher(row.teacher_name, users, fallback_user)
            match = self._match_lesson(row, schedule_index, summary)
            if match is None:
                continue
            if options['dry_run']:
                summary.rows_imported += 1
                continue
            try:
                self._import_row(row, match, actor, summary)
            except Exception as exc:  # noqa: BLE001 - command must continue and report row-level failures.
                summary.row_errors.append(f'{row.sheet_name} linha {row.row_number}: {exc}')

        self._print_summary(summary, dry_run=options['dry_run'])

    def _student_lookup(self):
        students = list(PedagogicalStudent.objects.all())
        exact = {_normalize(student.name): student for student in students}

        def find(sheet_name):
            normalized = _normalize(_sheet_student_name(sheet_name))
            if not normalized:
                return None
            if normalized in exact:
                return exact[normalized]
            matches = [
                student for student in students
                if _normalize(student.name).startswith(normalized) or normalized.startswith(_normalize(student.name))
            ]
            return matches[0] if len(matches) == 1 else None

        return find

    def _user_lookup(self):
        users = list(get_user_model().objects.all())
        normalized_users = []
        for user in users:
            labels = [user.get_full_name(), user.username]
            normalized_users.append((user, [_normalize(label) for label in labels if label]))

        def find(teacher_name):
            normalized = _normalize(teacher_name)
            if not normalized:
                return None
            matches = [
                user for user, labels in normalized_users
                if any(label == normalized or label.startswith(normalized) or normalized.startswith(label) for label in labels)
            ]
            return matches[0] if len(matches) == 1 else None

        return find

    def _fallback_user(self, username):
        username = (username or '').strip()
        if not username:
            return None
        try:
            return get_user_model().objects.get(username=username)
        except get_user_model().DoesNotExist as exc:
            raise CommandError(f'Usuário fallback não encontrado: {username}') from exc

    def _parse_workbook(self, workbook, students, start_date, end_date, summary):
        rows = []
        for worksheet in workbook.worksheets:
            if _normalize(worksheet.title) == 'modelo':
                continue
            student = students(worksheet.title)
            header_row, header_keys = self._find_header(worksheet)
            if not header_row:
                continue
            columns = self._columns(header_keys)
            last_position = {'module': None, 'lesson': None}
            for row_number, raw_row in enumerate(
                worksheet.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                class_date = _parse_date(_value(raw_row, columns['date']))
                if not class_date or class_date < start_date or class_date > end_date:
                    continue
                summary.rows_in_period += 1
                if student is None:
                    summary.skipped_unmatched_student += 1
                    continue
                parsed = self._parse_feedback_row(
                    worksheet.title,
                    row_number,
                    raw_row,
                    columns,
                    student,
                    class_date,
                    last_position,
                )
                if parsed is None:
                    summary.skipped_incomplete += 1
                    continue
                last_position['module'] = parsed.module_number
                last_position['lesson'] = parsed.lesson_number
                rows.append(parsed)
        return rows

    def _find_header(self, worksheet):
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=min(10, worksheet.max_row), values_only=True),
            start=1,
        ):
            keys = [_normalize(value) for value in row]
            if 'data' in keys and 'montagem' in keys:
                return row_number, keys
        return None, []

    def _columns(self, header_keys):
        return {
            'module': _header_index(header_keys, 'Módulo', 'Modulo'),
            'lesson_candidates': _header_indices(header_keys, 'Aula', 'Aulas'),
            'course': _header_index(header_keys, 'Curso'),
            'date': _header_index(header_keys, 'Data'),
            'teacher': _header_index(header_keys, 'Professor'),
            'punctuality': (
                _header_index(header_keys, 'Nota Pontualidade')
                or _header_index(header_keys, 'Pontualidade')
            ),
            'assembly_comment': _header_index(header_keys, 'Montagem'),
            'assembly_score': _header_index(header_keys, 'Montagem Nota'),
            'programming_comment': _header_index(header_keys, 'Programação', 'Programacao'),
            'programming_score': _header_index(header_keys, 'Programação Nota', 'Programacao Nota'),
            'participation_comment': _header_index_contains(
                header_keys,
                include=('interesse', 'participacao'),
                exclude=('nota',),
            ),
            'participation_score': _header_index_contains(
                header_keys,
                include=('interesse', 'participacao', 'nota'),
            ),
            'behavior_comment': _header_index(header_keys, 'Comportamento'),
            'behavior_score': _header_index(header_keys, 'Comportamento Nota'),
            'general_comment': _header_index(header_keys, 'Comentário Geral', 'Comentario Geral'),
        }

    def _parse_feedback_row(self, sheet_name, row_number, raw_row, columns, student, class_date, last_position):
        module_number = _parse_int(_value(raw_row, columns['module']))
        lesson_number = self._lesson_number(raw_row, columns['lesson_candidates'])
        if module_number is None:
            module_number = self._module_from_lesson_or_sequence(lesson_number, last_position)
        if lesson_number is not None and lesson_number > 15:
            module_number = ((lesson_number - 1) // 15) + 1
            lesson_number = ((lesson_number - 1) % 15) + 1
        if lesson_number is None and last_position['lesson']:
            lesson_number = last_position['lesson'] + 1
            if lesson_number > 15:
                module_number = (last_position['module'] or 1) + 1
                lesson_number = 1
        punctuality_score = _parse_punctuality(_value(raw_row, columns['punctuality']))
        assembly_score = _parse_score(_value(raw_row, columns['assembly_score']))
        programming_score = _parse_score(_value(raw_row, columns['programming_score']))
        participation_score = _parse_score(_value(raw_row, columns['participation_score']))
        behavior_score = _parse_score(_value(raw_row, columns['behavior_score']))
        assembly_comment = _text(_value(raw_row, columns['assembly_comment']))
        programming_comment = _text(_value(raw_row, columns['programming_comment']))
        participation_comment = _text(_value(raw_row, columns['participation_comment']))
        behavior_comment = _text(_value(raw_row, columns['behavior_comment']))
        general_comment = _text(_value(raw_row, columns['general_comment']))
        has_programming = not _is_programming_empty(programming_comment, programming_score)

        if (
            module_number not in {1, 2, 3}
            or not lesson_number
            or lesson_number < 1
            or lesson_number > 15
            or punctuality_score is None
            or assembly_score is None
            or participation_score is None
            or behavior_score is None
            or not assembly_comment
            or not participation_comment
            or not behavior_comment
            or not general_comment
        ):
            return None
        if has_programming and (programming_score is None or not programming_comment):
            return None

        return FeedbackImportRow(
            sheet_name=sheet_name,
            row_number=row_number,
            student=student,
            class_date=class_date,
            module_number=module_number,
            lesson_number=lesson_number,
            course_hint=_text(_value(raw_row, columns['course'])),
            teacher_name=_text(_value(raw_row, columns['teacher'])),
            punctuality_score=punctuality_score,
            assembly_comment=assembly_comment,
            assembly_score=assembly_score,
            has_programming=has_programming,
            programming_comment=programming_comment if has_programming else '',
            programming_score=programming_score if has_programming else None,
            participation_comment=participation_comment,
            participation_score=participation_score,
            behavior_comment=behavior_comment,
            behavior_score=behavior_score,
            general_comment=general_comment,
        )

    def _lesson_number(self, raw_row, candidate_indices):
        for index in candidate_indices:
            lesson_number = _parse_int(_value(raw_row, index))
            if lesson_number is not None:
                return lesson_number
        return None

    def _module_from_lesson_or_sequence(self, lesson_number, last_position):
        if lesson_number and lesson_number > 15:
            return ((lesson_number - 1) // 15) + 1
        if last_position['module']:
            return last_position['module']
        return 1

    def _existing_lesson_index(self, rows):
        student_ids = {row.student.id for row in rows}
        dates = {row.class_date for row in rows}
        index = {}
        if not student_ids or not dates:
            return index
        lessons = (
            Lesson.objects
            .filter(student_id__in=student_ids, date__in=dates, lesson_type=Lesson.TYPE_REGULAR)
            .select_related('student', 'course', 'room')
        )
        for lesson in lessons:
            index.setdefault((lesson.student_id, lesson.date), []).append(ScheduleMatch(record=None, lesson=lesson))
        return index

    def _load_sponte_schedule(self, rows, schedule_index, start_date, end_date, summary, options):
        needed_dates_by_student = {}
        for row in rows:
            needed_dates_by_student.setdefault(row.student.id, set()).add(row.class_date)
        students = {
            row.student.id: row.student
            for row in rows
            if row.student.external_id
        }
        client = SponteSOAPClient()
        calls = 0
        for student_id, student in sorted(students.items(), key=lambda item: item[1].name):
            for chunk_start, chunk_end in _date_chunks(start_date, end_date):
                if not any(chunk_start <= day <= chunk_end for day in needed_dates_by_student.get(student_id, set())):
                    continue
                try:
                    response = client.call(
                        'GetAgendaAluno',
                        {
                            'nAlunoID': student.external_id,
                            'sTurmaID': '',
                            'sCursoID': '',
                            'dDataInicio': chunk_start.strftime('%d/%m/%Y'),
                            'dDataTermino': chunk_end.strftime('%d/%m/%Y'),
                        },
                        use_cache=True,
                    )
                    records = parse_sponte_free_class_schedule(
                        response.xml_text,
                        student_external_id=student.external_id,
                    )
                except SponteAPIError as exc:
                    summary.api_errors.append(f'{student.name}: {exc}')
                    continue

                calls += 1
                for record in records:
                    class_date = _parse_sponte_date(record.get('date'))
                    start_time = _parse_sponte_time(record.get('start_time'))
                    end_time = _parse_sponte_time(record.get('end_time'))
                    free_class_id = (record.get('free_class_id') or '').strip()
                    if not all([class_date, start_time, end_time, free_class_id]):
                        continue
                    if class_date not in needed_dates_by_student.get(student_id, set()):
                        continue
                    schedule_index.setdefault((student_id, class_date), []).append(ScheduleMatch(record=record))

                if options['pause_every'] > 0 and calls % options['pause_every'] == 0:
                    time.sleep(max(options['pause_seconds'], 0))

    def _actor_for_teacher(self, teacher_name, users, fallback_user):
        return users(teacher_name) or fallback_user

    def _match_lesson(self, row, schedule_index, summary):
        candidates = schedule_index.get((row.student.id, row.class_date), [])
        if not candidates:
            summary.skipped_no_schedule += 1
            return None

        seen = set()
        unique_candidates = []
        for candidate in candidates:
            key = self._candidate_key(candidate)
            if key in seen:
                continue
            seen.add(key)
            unique_candidates.append(candidate)

        if len(unique_candidates) == 1:
            return unique_candidates[0]

        feedback_matches = []
        for candidate in unique_candidates:
            if not candidate.lesson:
                continue
            try:
                feedback = candidate.lesson.feedback
            except LessonFeedback.DoesNotExist:
                continue
            if feedback.module_number == row.module_number and feedback.lesson_number == row.lesson_number:
                feedback_matches.append(candidate)
        if len(feedback_matches) == 1:
            return feedback_matches[0]

        record_candidates = [candidate for candidate in unique_candidates if candidate.record]
        if len(record_candidates) == 1:
            return record_candidates[0]

        course_hint = _normalize(row.course_hint)
        if course_hint:
            course_matches = []
            for candidate in unique_candidates:
                course_name = ''
                if candidate.record:
                    course_name = candidate.record.get('course_name') or ''
                elif candidate.lesson and candidate.lesson.course_id:
                    course_name = candidate.lesson.course.name
                normalized_course = _normalize(course_name)
                if normalized_course and (normalized_course.startswith(course_hint) or course_hint.startswith(normalized_course)):
                    course_matches.append(candidate)
            if len(course_matches) == 1:
                return course_matches[0]

        summary.skipped_ambiguous_schedule += 1
        return None

    def _candidate_key(self, candidate):
        if candidate.lesson:
            return ('lesson', candidate.lesson.id)
        record = candidate.record or {}
        return (
            'record',
            record.get('student_external_id'),
            record.get('free_class_id'),
            record.get('date'),
            record.get('start_time'),
            record.get('end_time'),
        )

    def _import_row(self, row, match, actor, summary):
        with transaction.atomic():
            lesson, lesson_created, lesson_updated = self._upsert_lesson(row, match, actor)
            feedback, feedback_created, feedback_updated = self._upsert_feedback(row, lesson, actor)
            report_task, report_created = create_pedagogical_report_task_for_feedback(feedback, actor=actor)
            opportunity, opportunity_created = create_post_sale_opportunity_for_lesson_feedback(feedback, actor=actor)

        summary.rows_imported += 1
        summary.lessons_created += int(lesson_created)
        summary.lessons_updated += int(lesson_updated)
        summary.feedbacks_created += int(feedback_created)
        summary.feedbacks_updated += int(feedback_updated)
        summary.feedbacks_unchanged += int(not feedback_created and not feedback_updated)
        summary.report_tasks_created += int(bool(report_task and report_created))
        summary.post_sale_opportunities_created += int(bool(opportunity and opportunity_created))

    def _upsert_lesson(self, row, match, actor):
        now = timezone.now()
        if match.lesson:
            values = {'status': Lesson.STATUS_DONE}
            changed = self._update_instance(match.lesson, values)
            if changed:
                match.lesson.save()
            return match.lesson, False, changed

        record = match.record or {}
        class_date = _parse_sponte_date(record.get('date'))
        start_time = _parse_sponte_time(record.get('start_time'))
        end_time = _parse_sponte_time(record.get('end_time'))
        free_class_id = (record.get('free_class_id') or '').strip()
        base_external_id = f'aula_livre:{row.student.external_id}:{free_class_id}:{class_date.isoformat()}'
        external_id = base_external_id
        course = _course_for_schedule(record)
        room = _room_for_schedule(record)
        notes = '\n'.join([
            'Aula histórica marcada como Presença a partir da planilha de feedback.',
            'Horário obtido do Sponte para vincular o feedback.',
            f'Professor: {record.get("teacher_name") or row.teacher_name or "-"}',
            f'CursoID: {record.get("course_external_id") or "-"}',
            f'DisciplinaID: {record.get("discipline_external_id") or "-"}',
        ])
        values = {
            'student': row.student,
            'student_name_snapshot': row.student.name,
            'responsible_name_snapshot': row.student.responsible_name,
            'whatsapp_snapshot': row.student.whatsapp,
            'lesson_type': Lesson.TYPE_REGULAR,
            'course': course,
            'room': room,
            'date': class_date,
            'start_time': start_time,
            'end_time': end_time,
            'status': Lesson.STATUS_DONE,
            'notes': notes,
            'source': Lesson.SOURCE_SPONTE,
            'external_id': external_id,
            'synced_at': now,
        }
        lesson = Lesson.objects.filter(source=Lesson.SOURCE_SPONTE, external_id=external_id).first()
        if lesson is not None and self._lesson_has_other_feedback(lesson, row):
            external_id = f'{base_external_id}:feedback:{row.module_number}-{row.lesson_number}'
            values['external_id'] = external_id
            lesson = Lesson.objects.filter(source=Lesson.SOURCE_SPONTE, external_id=external_id).first()
        if lesson is None:
            values['created_by'] = actor
            return Lesson.objects.create(**values), True, False
        changed = self._update_instance(lesson, values)
        if changed:
            lesson.save()
        return lesson, False, changed

    def _lesson_has_other_feedback(self, lesson, row):
        try:
            feedback = lesson.feedback
        except LessonFeedback.DoesNotExist:
            return False
        return feedback.module_number != row.module_number or feedback.lesson_number != row.lesson_number

    def _upsert_feedback(self, row, lesson, actor):
        values = {
            'module_number': row.module_number,
            'lesson_number': row.lesson_number,
            'punctuality_score': row.punctuality_score,
            'assembly_comment': row.assembly_comment,
            'assembly_score': row.assembly_score,
            'has_programming': row.has_programming,
            'programming_comment': row.programming_comment,
            'programming_score': row.programming_score,
            'participation_comment': row.participation_comment,
            'participation_score': row.participation_score,
            'behavior_comment': row.behavior_comment,
            'behavior_score': row.behavior_score,
            'general_comment': row.general_comment,
            'updated_by': actor,
        }
        feedback = LessonFeedback.objects.filter(lesson=lesson).first()
        if feedback is None:
            values['created_by'] = actor
            feedback = LessonFeedback(lesson=lesson, **values)
            feedback.save()
            return feedback, True, False
        changed = self._update_instance(feedback, values)
        if changed:
            feedback.save()
        return feedback, False, changed

    def _update_instance(self, instance, values):
        changed = False
        for field_name, value in values.items():
            if getattr(instance, field_name) != value:
                setattr(instance, field_name, value)
                changed = True
        return changed

    def _print_summary(self, summary, *, dry_run):
        label = 'Diagnóstico' if dry_run else 'Importação'
        self.stdout.write(self.style.SUCCESS(f'{label} finalizado.'))
        for name, value in [
            ('rows_in_period', summary.rows_in_period),
            ('rows_ready', summary.rows_ready),
            ('rows_imported', summary.rows_imported),
            ('skipped_unmatched_student', summary.skipped_unmatched_student),
            ('skipped_incomplete', summary.skipped_incomplete),
            ('skipped_no_schedule', summary.skipped_no_schedule),
            ('skipped_ambiguous_schedule', summary.skipped_ambiguous_schedule),
            ('lessons_created', summary.lessons_created),
            ('lessons_updated', summary.lessons_updated),
            ('feedbacks_created', summary.feedbacks_created),
            ('feedbacks_updated', summary.feedbacks_updated),
            ('feedbacks_unchanged', summary.feedbacks_unchanged),
            ('report_tasks_created', summary.report_tasks_created),
            ('post_sale_opportunities_created', summary.post_sale_opportunities_created),
        ]:
            self.stdout.write(f'{name}: {value}')
        for error in summary.api_errors[:20]:
            self.stdout.write(self.style.WARNING(f'API: {error}'))
        for error in summary.row_errors[:20]:
            self.stdout.write(self.style.WARNING(f'Linha: {error}'))
        if len(summary.api_errors) > 20 or len(summary.row_errors) > 20:
            self.stdout.write(self.style.WARNING('Há mais erros omitidos no resumo.'))
